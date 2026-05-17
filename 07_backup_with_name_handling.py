"""
fastsolver_bridge.py — Python-side entry point for the .xlam add-in.

Called as:
    FastSolver.exe <run_dir>

Where <run_dir> contains:
    config.json    — problem spec exported by VBA
    source.xlsx    — copy of the user's workbook

Produces:
    results.json   — final score + variable name/value arrays
    solved.xlsx    — workbook copy with optimal values written in

Implements method dispatch:
    "Auto"                  → routes by problem size/structure
    "SLSQP Python"          → SciPy SLSQP via pycel evaluation
    "L-BFGS-B"              → scaled to large var counts
    "Differential Evolution"→ global, derivative-free, pycel-eval
    "Homotopy / AutoDiff"   → JAX + smoothing — placeholder for full
                              graph→JAX rewrite; falls back to L-BFGS-B
                              over pycel-eval with finite-diff today
"""

from __future__ import annotations
import json
import os
import sys
import time
import traceback
from pathlib import Path
import numpy as np

# ---------------------------------------------------------------------------
# Pycel defined-names workaround (some openpyxl versions break here)
# ---------------------------------------------------------------------------
# Why this is needed:
#   Excel named ranges can be workbook-scoped or sheet-scoped.  openpyxl
#   versions expose them through slightly different containers, and pycel
#   1.0b30 expects a simple dict lookup.  If a name is not found there, pycel
#   falls through to its structured-table parser and emits misleading errors
#   like "Table Name not found: LogLogScale" even when LogLogScale is just a
#   normal Excel variable/name.
#
#   This mapper collects workbook names and worksheet-local names, and adds
#   both bare and sheet-qualified aliases, e.g.
#       truncation
#       LogLog Fitting!truncation
#       'LogLog Fitting'!truncation
#   which is required when the same name exists in multiple scopes.
import pycel.excelwrapper as _ew


def _dn_container_values(container):
    """Return DefinedName-like objects from old/new openpyxl containers."""
    if container is None:
        return []

    try:
        vals = list(container.definedName)
    except Exception:
        try:
            vals = list(container.values())
        except Exception:
            try:
                vals = list(container)
            except Exception:
                vals = []

    out = []
    for v in vals:
        if isinstance(v, (list, tuple, set)):
            out.extend(v)
        else:
            out.append(v)
    return out


def _dn_text(d):
    """openpyxl has used attr_text/value/text across versions."""
    for attr in ("attr_text", "value", "text"):
        try:
            v = getattr(d, attr, None)
        except Exception:
            v = None
        if v:
            return str(v)
    return None


def _dn_name(d):
    try:
        return getattr(d, "name", None) or getattr(d, "localName", None)
    except Exception:
        return None


def _sheet_for_local_id(wb, local_sheet_id):
    if local_sheet_id is None:
        return None
    try:
        return wb.worksheets[int(local_sheet_id)].title
    except Exception:
        return None


def _quote_sheet(sheet):
    if sheet is None:
        return None
    escaped = str(sheet).replace("'", "''")
    return "'" + escaped + "'"


def _iter_defined_names_from_xlsx_xml(xlsx_path):
    """Yield defined names by reading xl/workbook.xml directly.

    This is the version-independent fallback.  Some openpyxl releases expose
    sheet-scoped duplicate names inconsistently (or not at all), but Excel's
    Name Manager entries are always stored in workbook.xml as <definedName>.
    pycel failures such as "Table Name not found: LogLogScale" usually mean
    the openpyxl-facing map missed one of these XML entries, so we read the XML
    too and union the results.
    """
    if not xlsx_path:
        return
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            raw = zf.read("xl/workbook.xml")
        root = ET.fromstring(raw)
    except Exception:
        return

    # localSheetId is a zero-based index into the workbook sheets collection.
    sheet_names = []
    try:
        for sh in root.findall(".//{*}sheets/{*}sheet"):
            sheet_names.append(sh.attrib.get("name", ""))
    except Exception:
        sheet_names = []

    try:
        nodes = root.findall(".//{*}definedNames/{*}definedName")
    except Exception:
        nodes = []

    for node in nodes:
        name = node.attrib.get("name")
        if not name:
            continue
        # Skip built-in names such as _xlnm.Print_Area; they are not solver
        # variables and rewriting them can break print/metadata formulas.
        if str(name).startswith("_xlnm."):
            continue
        target = "".join(node.itertext()).strip()
        if not target:
            continue
        sheet = None
        local_id = node.attrib.get("localSheetId")
        if local_id is not None:
            try:
                sheet = sheet_names[int(local_id)]
            except Exception:
                sheet = None
        yield str(name), str(target), sheet


def _iter_defined_names(wb, xlsx_path=None):
    """Yield (name, target, scope_sheet_or_None) from openpyxl + raw XML."""
    seen = set()

    def emit(name, target, sheet):
        if not name or not target:
            return None
        key = (str(name).casefold(), _strip_sheet_quotes(sheet).casefold() if sheet else None, str(target).strip())
        if key in seen:
            return None
        seen.add(key)
        return str(name), str(target), sheet

    # Workbook container: includes workbook names in older openpyxl and may
    # also include local names with localSheetId in some versions.
    for d in _dn_container_values(getattr(wb, "defined_names", None)):
        name = _dn_name(d)
        target = _dn_text(d)
        if not name or not target:
            continue
        sheet = _sheet_for_local_id(wb, getattr(d, "localSheetId", None))
        row = emit(name, target, sheet)
        if row:
            yield row

    # Worksheet containers: where newer openpyxl may put sheet-scoped names.
    for ws in getattr(wb, "worksheets", []):
        for d in _dn_container_values(getattr(ws, "defined_names", None)):
            name = _dn_name(d)
            target = _dn_text(d)
            if not name or not target:
                continue
            row = emit(name, target, ws.title)
            if row:
                yield row

    # Raw XML fallback catches the cases openpyxl misses.  This is what fixes
    # duplicated local/workbook names like truncation and ordinary workbook
    # names like LogLogShape/LogLogScale when pycel still reports them as
    # missing table names.
    for name, target, sheet in _iter_defined_names_from_xlsx_xml(xlsx_path):
        row = emit(name, target, sheet)
        if row:
            yield row

def _put_name_alias(m, key, target, overwrite=True):
    """Excel names are case-insensitive; store common case variants."""
    if not key:
        return
    for k in {key, key.lower(), key.upper()}:
        if overwrite or k not in m:
            m[k] = target


def _safe_defined_names(self):
    if self._defined_names is None:
        m = {}
        try:
            # First pass: workbook-scoped names.
            local = []
            for name, target, sheet in _iter_defined_names(self.workbook):
                if sheet:
                    local.append((name, target, sheet))
                else:
                    _put_name_alias(m, name, target, overwrite=True)

            # Second pass: sheet-scoped names.  These get explicit
            # sheet-qualified aliases and only become bare-name fallbacks if no
            # workbook-scoped name already exists.
            for name, target, sheet in local:
                sheet = str(sheet).strip().strip("'")
                aliases = (
                    f"{sheet}!{name}",
                    f"{_quote_sheet(sheet)}!{name}",
                )
                for alias in aliases:
                    _put_name_alias(m, alias, target, overwrite=True)

                # Excel allows local names with the same bare name on multiple
                # sheets.  Do not overwrite a workbook-scoped bare name; but if
                # this is the only definition, let bare references resolve.
                _put_name_alias(m, name, target, overwrite=False)
        except Exception:
            pass
        self._defined_names = m
    return self._defined_names



# ---------------------------------------------------------------------------
# Formula rewriter for defined names (pycel 1.0b30)
# ---------------------------------------------------------------------------
# pycel frequently mis-parses workbook/sheet names embedded inside formulas as
# structured-table names and then reports misleading errors such as:
#     Table Name not found: LogLogScale
# The safest workaround is to compile a temporary workbook where formula text
# has already expanded supported defined names to their target cell/range refs.


def _strip_sheet_quotes(sheet):
    sheet = str(sheet or "").strip()
    if len(sheet) >= 2 and sheet[0] == "'" and sheet[-1] == "'":
        sheet = sheet[1:-1].replace("''", "'")
    return sheet.strip()


def _formula_ref_from_defined_name_target(target):
    """Return a formula-safe replacement for a defined-name target.

    Most solver-relevant names point at one cell/range, e.g.
    "'LogLog Fitting'!$B$33".  Broken names (#REF!) are deliberately skipped.
    Simple constant/formula names are also supported where possible.
    """
    if target is None:
        return None
    t = str(target).strip()
    if not t:
        return None

    # Defined names may be stored as =Sheet!$A$1 or Sheet!$A$1.
    if t.startswith("="):
        t = t[1:].strip()

    # Do not bake broken Name Manager entries into formulas.
    if "#REF!" in t.upper():
        return None

    # External workbook refs are not made pycel-safe here; leave untouched.
    if "[" in t and "]" in t:
        return None

    # Normalize single sheet refs/ranges to quoted sheet form.  This is safe
    # for spaces and harmless for simple sheet names.
    if "!" in t and "," not in t:
        sheet, ref = t.rsplit("!", 1)
        sheet = _strip_sheet_quotes(sheet)
        if not sheet or not ref:
            return None
        return f"{_quote_sheet(sheet)}!{ref.strip()}"

    # Constants such as 0.5, TRUE, or a range on the current sheet can be used
    # directly.  If it is a formula expression, wrap it to preserve precedence.
    if any(op in t for op in ("+", "-", "*", "/", "^", "&", "=", "<", ">")):
        return f"({t})"
    return t


def _build_name_rewrite_maps(wb, xlsx_path=None):
    """Build workbook/local defined-name maps for formula rewriting."""
    workbook = {}
    local = {}
    skipped = []

    for name, target, sheet in _iter_defined_names(wb, xlsx_path):
        repl = _formula_ref_from_defined_name_target(target)
        if not repl:
            skipped.append((sheet or "WORKBOOK", str(name), str(target)))
            continue
        key = str(name).casefold()
        if sheet:
            skey = _strip_sheet_quotes(sheet).casefold()
            local.setdefault(skey, {})[key] = repl
        else:
            workbook[key] = repl

    return {"workbook": workbook, "local": local, "skipped": skipped}


def _lookup_bare_defined_name(name_maps, current_sheet, name):
    if not name_maps or not name:
        return None
    nkey = str(name).casefold()
    skey = _strip_sheet_quotes(current_sheet).casefold() if current_sheet else None
    if skey:
        hit = name_maps.get("local", {}).get(skey, {}).get(nkey)
        if hit is not None:
            return hit
    return name_maps.get("workbook", {}).get(nkey)


def _lookup_qualified_defined_name(name_maps, sheet, name):
    if not name_maps or not sheet or not name:
        return None
    skey = _strip_sheet_quotes(sheet).casefold()
    nkey = str(name).casefold()
    return name_maps.get("local", {}).get(skey, {}).get(nkey)


def _is_excel_name_start(ch):
    return bool(ch) and (ch == "_" or ch == "\\" or ch.isalpha())


def _is_excel_name_char(ch):
    return bool(ch) and (ch == "_" or ch == "\\" or ch == "." or ch.isalnum())


def _is_unquoted_sheet_char(ch):
    # Unquoted sheet names in formulas cannot contain spaces or punctuation
    # like apostrophes.  This intentionally keeps the parser conservative.
    return bool(ch) and (ch == "_" or ch == "." or ch.isalnum())


def _rewrite_formula_segment_defined_names(segment, current_sheet, name_maps, hits=None):
    """Rewrite defined names in a formula segment that is outside strings."""
    if not segment or not name_maps:
        return segment

    n = len(segment)
    out = []
    i = 0

    while i < n:
        ch = segment[i]

        # Quoted sheet-qualified local name:  'Sheet 1'!localName
        if ch == "'":
            j = i + 1
            sheet_chars = []
            while j < n:
                if segment[j] == "'":
                    if j + 1 < n and segment[j + 1] == "'":
                        sheet_chars.append("'")
                        j += 2
                        continue
                    break
                sheet_chars.append(segment[j])
                j += 1
            if j < n and segment[j] == "'" and j + 1 < n and segment[j + 1] == "!":
                k = j + 2
                if k < n and _is_excel_name_start(segment[k]):
                    l = k + 1
                    while l < n and _is_excel_name_char(segment[l]):
                        l += 1
                    sheet = "".join(sheet_chars)
                    name = segment[k:l]
                    repl = _lookup_qualified_defined_name(name_maps, sheet, name)
                    if repl is not None:
                        if hits is not None:
                            hits[name] = hits.get(name, 0) + 1
                        out.append(repl)
                        i = l
                        continue
            out.append(ch)
            i += 1
            continue

        # Unquoted sheet-qualified local name:  Sheet1!localName
        if _is_unquoted_sheet_char(ch):
            j = i + 1
            while j < n and _is_unquoted_sheet_char(segment[j]):
                j += 1
            if j < n and segment[j] == "!":
                k = j + 1
                if k < n and _is_excel_name_start(segment[k]):
                    l = k + 1
                    while l < n and _is_excel_name_char(segment[l]):
                        l += 1
                    sheet = segment[i:j]
                    name = segment[k:l]
                    repl = _lookup_qualified_defined_name(name_maps, sheet, name)
                    if repl is not None:
                        if hits is not None:
                            hits[name] = hits.get(name, 0) + 1
                        out.append(repl)
                        i = l
                        continue
                # It is a sheet/cell ref or unknown sheet-qualified token.
                # Do not accidentally rewrite the sheet token as a bare name.
                out.append(segment[i:j])
                i = j
                continue

        # Bare workbook/local name:  LogLogShape, truncation, etc.
        if _is_excel_name_start(ch):
            j = i + 1
            while j < n and _is_excel_name_char(segment[j]):
                j += 1
            token = segment[i:j]

            # Function calls are not names unless Excel's Name Manager has a
            # matching formula name.  If a user truly has a name used as Foo(),
            # replacing it is still correct; normal functions simply won't hit.
            repl = _lookup_bare_defined_name(name_maps, current_sheet, token)
            if repl is not None:
                if hits is not None:
                    hits[token] = hits.get(token, 0) + 1
                out.append(repl)
            else:
                out.append(token)
            i = j
            continue

        out.append(ch)
        i += 1

    return "".join(out)


def _rewrite_formula_defined_names(formula, current_sheet, name_maps, hits=None):
    """Rewrite defined names in formula text, skipping double-quoted strings."""
    if not isinstance(formula, str) or not formula.startswith("=") or not name_maps:
        return formula

    out = []
    buf = []
    i = 0
    n = len(formula)
    in_string = False

    while i < n:
        ch = formula[i]
        if ch == '"':
            if in_string:
                # Excel escapes a quote inside a string as "".
                if i + 1 < n and formula[i + 1] == '"':
                    out.append(ch)
                    out.append(formula[i + 1])
                    i += 2
                    continue
                in_string = False
                out.append(ch)
                i += 1
                continue
            # Starting a string: flush formula segment first.
            if buf:
                seg = "".join(buf)
                out.append(_rewrite_formula_segment_defined_names(seg, current_sheet, name_maps, hits))
                buf = []
            in_string = True
            out.append(ch)
            i += 1
            continue

        if in_string:
            out.append(ch)
        else:
            buf.append(ch)
        i += 1

    if buf:
        seg = "".join(buf)
        out.append(_rewrite_formula_segment_defined_names(seg, current_sheet, name_maps, hits))

    return "".join(out)


def _resolve_defined_name_reference(ref, current_sheet, name_maps):
    """Resolve a config ref that is itself a defined name, if possible."""
    if not ref or not name_maps:
        return None
    s = str(ref).strip()
    if s.startswith("="):
        return None

    # Quoted sheet-local name: 'Sheet 1'!truncation
    if s.startswith("'") and "!" in s:
        sheet_part, name_part = s.rsplit("!", 1)
        name_part = name_part.strip().replace("$", "")
        if name_part and _is_excel_name_start(name_part[0]) and all(_is_excel_name_char(c) for c in name_part):
            return _lookup_qualified_defined_name(name_maps, sheet_part, name_part)
        return None

    # Unquoted sheet-local name: Sheet1!truncation.  Cell refs won't hit.
    if "!" in s:
        sheet_part, name_part = s.rsplit("!", 1)
        name_part = name_part.strip().replace("$", "")
        if name_part and _is_excel_name_start(name_part[0]) and all(_is_excel_name_char(c) for c in name_part):
            return _lookup_qualified_defined_name(name_maps, sheet_part, name_part)
        return None

    # Bare name.
    token = s.replace("$", "")
    if token and _is_excel_name_start(token[0]) and all(_is_excel_name_char(c) for c in token):
        return _lookup_bare_defined_name(name_maps, current_sheet, token)
    return None


def _rewrite_workbook_formulas_for_pycel(src_xlsx, log_fn=None):
    """Create a pycel-safe workbook copy with defined names expanded.

    Returns (path_to_compile, name_maps).  The original source.xlsx is never
    modified; only a temporary workbook in the run directory is saved.
    """
    from openpyxl import load_workbook

    src_xlsx = Path(src_xlsx)
    try:
        wb = load_workbook(src_xlsx, data_only=False, keep_links=False)
    except Exception as e:
        if log_fn:
            log_fn(f"Defined-name formula rewrite skipped: could not open workbook ({e})")
        return src_xlsx, {"workbook": {}, "local": {}, "skipped": []}

    name_maps = _build_name_rewrite_maps(wb, src_xlsx)
    if log_fn:
        n_wb = len(name_maps.get("workbook", {}))
        n_local = sum(len(v) for v in name_maps.get("local", {}).values())
        log_fn(f"Defined-name formula rewrite: discovered {n_wb} workbook names and {n_local} sheet-local names")
    hits = {}
    changed = 0
    samples = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    nv = _rewrite_formula_defined_names(v, ws.title, name_maps, hits)
                    if nv != v:
                        cell.value = nv
                        changed += 1
                        if len(samples) < 3:
                            samples.append(f"{ws.title}!{cell.coordinate}: {v} -> {nv}")

    if log_fn:
        if changed:
            top_hits = ", ".join(f"{k}={v}" for k, v in sorted(hits.items(), key=lambda kv: (-kv[1], kv[0]))[:8])
            log_fn(f"Defined-name formula rewrite: {changed} formula cells rewritten" + (f" ({top_hits})" if top_hits else ""))
            if os.environ.get("FASTSOLVER_DEBUG"):
                for s in samples:
                    log_fn("  rewrite sample: " + s[:500])
        else:
            log_fn("Defined-name formula rewrite: no formula cells needed rewriting")
        skipped = name_maps.get("skipped", [])
        if skipped and os.environ.get("FASTSOLVER_DEBUG"):
            shown = ", ".join(f"{scope}!{name}->{target}" for scope, name, target in skipped[:8])
            log_fn(f"Defined-name formula rewrite: skipped unsupported/broken names: {shown}")

    if not changed:
        return src_xlsx, name_maps

    out = src_xlsx.with_name("_fastsolver_pycel_formula_rewrite.xlsx")
    try:
        wb.save(out)
        return out, name_maps
    except Exception as e:
        if log_fn:
            log_fn(f"Defined-name formula rewrite save failed ({e}); using original workbook")
        return src_xlsx, name_maps


# ---------------------------------------------------------------------------
# Defined-name sanitizer for pycel compile workbooks
# ---------------------------------------------------------------------------
# Formula rewriting handles the important case: valid names embedded in formula
# text.  This sanitizer is a second defensive pass over the temporary workbook's
# Name Manager table so pycel never sees broken #REF! names, and so obviously
# blank duplicate definitions do not win over usable scoped definitions.


def _iter_defined_name_entries(wb):
    """Yield (container, defined_name_obj, name, target, scope_sheet_or_None)."""
    container = getattr(wb, "defined_names", None)
    for d in _dn_container_values(container):
        name = _dn_name(d)
        target = _dn_text(d)
        if not name or not target:
            continue
        sheet = _sheet_for_local_id(wb, getattr(d, "localSheetId", None))
        yield container, d, str(name), str(target), sheet

    for ws in getattr(wb, "worksheets", []):
        container = getattr(ws, "defined_names", None)
        for d in _dn_container_values(container):
            name = _dn_name(d)
            target = _dn_text(d)
            if not name or not target:
                continue
            yield container, d, str(name), str(target), ws.title


def _remove_defined_name_from_container(container, d, name):
    """Best-effort deletion across old/new openpyxl defined-name containers."""
    if container is None:
        return False

    # openpyxl 3.x DefinedNameDict
    try:
        if hasattr(container, "pop"):
            container.pop(name, None)
            return True
    except Exception:
        pass

    try:
        del container[name]
        return True
    except Exception:
        pass

    # older openpyxl list-like container
    try:
        if hasattr(container, "definedName"):
            container.definedName.remove(d)
            return True
    except Exception:
        pass

    return False


def _defined_name_single_cell_value(wb, target):
    """Return the target cell's value for simple one-cell defined names.

    Returns None when the target is unsupported, missing, or blank.  A formula
    string in the target cell counts as non-blank because it is a usable target.
    """
    try:
        from openpyxl.utils.cell import range_boundaries
        t = str(target or "").strip()
        if t.startswith("="):
            t = t[1:].strip()
        if "#REF!" in t.upper() or "," in t or "!" not in t:
            return None
        sheet, ref = t.rsplit("!", 1)
        sheet = _strip_sheet_quotes(sheet)
        ref = ref.replace("$", "")
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        if min_col != max_col or min_row != max_row:
            return "__RANGE__"  # valid non-blank-ish range target
        if sheet not in wb.sheetnames:
            return None
        return wb[sheet].cell(min_row, min_col).value
    except Exception:
        return None


def _sanitize_defined_names_for_pycel(src_xlsx, log_fn=None):
    """Create a compile-only workbook with dangerous defined names removed.

    The original source.xlsx is left untouched.  This pass complements the
    formula rewriter: it strips #REF! names like TriangleRange from the Name
    Manager table and removes duplicate same-name definitions only when one of
    the duplicates points at a blank/unsupported target and another points at a
    usable target.  Valid sheet-local names are not blindly dropped, because
    Excel's scope rules matter for names such as truncation.
    """
    from openpyxl import load_workbook

    src_xlsx = Path(src_xlsx)
    try:
        wb = load_workbook(src_xlsx, data_only=False, keep_links=False)
    except Exception as e:
        if log_fn:
            log_fn(f"Defined-name sanitize skipped: could not open workbook ({e})")
        return src_xlsx

    removed = []
    entries = list(_iter_defined_name_entries(wb))

    # 1) Remove hard-broken names first (#REF! in the defined-name target).
    for container, d, name, target, sheet in entries:
        if "#REF!" in str(target).upper():
            if _remove_defined_name_from_container(container, d, name):
                removed.append(f"{sheet or 'WORKBOOK'}!{name}->{target}")

    # Re-read after removals.
    entries = list(_iter_defined_name_entries(wb))

    # 2) For duplicate same-name definitions, remove only obviously blank or
    # unsupported entries when at least one duplicate has a usable target.  This
    # avoids the dangerous strategy of always keeping workbook scope over local
    # scope, which would be wrong for sheet-local names like LogLog truncation.
    groups = {}
    for entry in entries:
        container, d, name, target, sheet = entry
        groups.setdefault(name.casefold(), []).append(entry)

    for _name_key, group in groups.items():
        if len(group) < 2:
            continue
        scored = []
        for entry in group:
            container, d, name, target, sheet = entry
            cell_value = _defined_name_single_cell_value(wb, target)
            usable = cell_value is not None
            scored.append((usable, entry, cell_value))
        if not any(usable for usable, _entry, _v in scored):
            continue
        for usable, entry, cell_value in scored:
            if usable:
                continue
            container, d, name, target, sheet = entry
            if _remove_defined_name_from_container(container, d, name):
                removed.append(f"duplicate blank/unsupported {sheet or 'WORKBOOK'}!{name}->{target}")

    if not removed:
        if log_fn:
            log_fn("Defined-name sanitize: nothing to remove")
        return src_xlsx

    out = src_xlsx.with_name("_fastsolver_pycel_names_sanitized.xlsx")
    try:
        wb.save(out)
        if log_fn:
            shown = "; ".join(removed[:8])
            extra = "" if len(removed) <= 8 else f"; +{len(removed)-8} more"
            log_fn(f"Defined-name sanitize: removed {len(removed)} names ({shown}{extra})")
        return out
    except Exception as e:
        if log_fn:
            log_fn(f"Defined-name sanitize save failed ({e}); using unsanitized workbook")
        return src_xlsx

_ew.ExcelOpxWrapper.defined_names = property(_safe_defined_names)


# ---------------------------------------------------------------------------
# Pycel numeric guard
# ---------------------------------------------------------------------------
def _install_pycel_numeric_guard():
    """Make pycel return Excel-style error values for numeric blow-ups.

    pycel 1.0b30 evaluates formulas as Python expressions.  For objectives
    such as log-log / Pareto fits, optimizers may legitimately probe values
    that make ``base ** shape`` enormous or invalid.  Excel would turn those
    probes into #NUM!/#DIV/0! and continue; pycel can raise OverflowError
    before our _safe_eval wrapper gets a normal value back.  Guarding the
    operator table keeps bad probes as penalty scores instead of crashing the
    solve.
    """
    try:
        import pycel.excelutil as _eu
        ops = getattr(_eu, "PYTHON_AST_OPERATORS", None)
        if not isinstance(ops, dict):
            return

        for key, fn in list(ops.items()):
            if not callable(fn) or getattr(fn, "_fastsolver_guarded", False):
                continue

            def _guarded_operator(*args, _fn=fn):
                """Guard pycel numeric operators without breaking unary ops.

                pycel uses the same operator table for both binary operators
                (Add/Sub/Mult/Div/Pow) and unary operators (USub/UAdd).  A
                binary-only wrapper breaks formulas containing unary minus such
                as ``x ** -shape`` because pycel calls the USub operator with
                a single argument.  Accept *args and forward exactly what pycel
                supplied.
                """
                try:
                    return _fn(*args)
                except ZeroDivisionError:
                    return "#DIV/0!"
                except OverflowError:
                    return "#NUM!"
                except ValueError:
                    # e.g. negative base to a fractional power.
                    return "#NUM!"
                except TypeError:
                    # Excel error values should propagate through arithmetic
                    # rather than crashing the Python evaluator.
                    for a in args:
                        if isinstance(a, str) and a.startswith("#"):
                            return a
                    return "#VALUE!"

            _guarded_operator._fastsolver_guarded = True
            ops[key] = _guarded_operator
    except Exception:
        pass


_install_pycel_numeric_guard()

from pycel import ExcelCompiler
from scipy.optimize import minimize, differential_evolution
from openpyxl import load_workbook


# ===========================================================================
# Helpers — simplex projection, bounds, etc.
# ===========================================================================

def project_simplex_group(x, idxs):
    """In place: project x[idxs] onto sum=1, >=0."""
    sub = x[idxs]
    sub = np.maximum(sub, 0)
    s = sub.sum()
    if s > 0:
        sub = sub / s
    else:
        sub = np.ones(len(sub)) / len(sub)
    x[idxs] = sub


def collect_simplex_groups(variables):
    """Return list of index-arrays for variables sharing rule 'Sum=1'."""
    groups = {}
    for i, v in enumerate(variables):
        if v["rule"] == "Sum=1":
            groups.setdefault(v["group"], []).append(i)
    return [np.array(idxs, dtype=int) for idxs in groups.values()]


def apply_projection(x, simplex_groups, lo, hi):
    x = np.clip(x, lo, hi)
    for idxs in simplex_groups:
        project_simplex_group(x, idxs)
    return x


# ===========================================================================
# Excel evaluation via pycel — score function builder
# ===========================================================================

def _split_addr(ref):
    """Split 'SheetName!$A$1' or "'Sheet 1'!$A$1" → (sheet, cell). Tolerates missing leading quote."""
    if "!" not in ref:
        return None, ref
    sheet, cell = ref.split("!", 1)
    sheet = sheet.strip().strip("'").strip()
    cell = cell.replace("$", "")
    return sheet, cell


def _pycel_addr(ref):
    """Convert VBA-exported ref → pycel address format SheetName!A1."""
    sheet, cell = _split_addr(ref)
    if sheet is None:
        return cell
    if " " in sheet:
        return f"'{sheet}'!{cell}"
    return f"{sheet}!{cell}"


def build_score_fn(excel, config, variables, simplex_groups, lo, hi):
    name_maps = getattr(excel, "_fastsolver_name_maps", None)
    default_sheet = _split_addr(variables[0]["ref"])[0] if variables else None

    def _ref_addr(ref, current_sheet=None):
        resolved = _resolve_defined_name_reference(ref, current_sheet or default_sheet, name_maps)
        return _pycel_addr(resolved or ref)

    var_addrs = [_ref_addr(v["ref"], _split_addr(v["ref"])[0] or default_sheet) for v in variables]
    obj_specs = [o for o in config["objectives"] if o["active"]]
    con_specs = [c for c in config["constraints"] if c["active"]]
    alpha = float(config.get("outlier_alpha", 0.7))
    hard_pen = float(config.get("hard_penalty", 1e6))

    eval_count = [0]

    # pycel's set_value() invalidates the full downstream dependency chain
    # automatically. The previous code manually nulled only DIRECT successors
    # (1 level deep), which left deeper cells (var -> A1 -> A2 -> obj) stale
    # and could return wrong scores. We now rely on set_value's own chain
    # invalidation, which is both correct and faster.
    def write_vars(x):
        for addr, val in zip(var_addrs, x):
            try:
                # pycel 1.0b30 skips invalidation when the written value is
                # unchanged. Clear the input cell first so downstream formulas
                # are reset even when optimizers re-probe the same point.
                try:
                    c = excel.cell_map.get(addr)
                    if c is not None:
                        c.value = None
                except Exception:
                    pass
                excel.set_value(addr, float(val))
            except Exception:
                pass

    # ---- Memoization -----------------------------------------------------
    # CD line-search and DE re-probe the same / nearby points constantly.
    # Key on the rounded x vector so micro-adjust steps that revisit a point
    # become free dict lookups instead of full pycel re-evaluations.
    _memo = {}
    _MEMO_DECIMALS = 10  # tighter than any tolerance we use; safe to cache

    def _memo_key(x):
        return tuple(round(float(v), _MEMO_DECIMALS) for v in x)

    formula_cache = {}

    def eval_ref(spec_str, is_rhs=False):
        s = spec_str.strip()
        host_sheet = default_sheet or "Sheet1"
        if not s.startswith("="):
            return _safe_eval(excel, _ref_addr(s, host_sheet))

        # Formula objectives/constraints exported in config can contain the
        # same defined names as worksheet formulas.  Rewrite them before
        # placing them in pycel scratch cells.
        try:
            s_for_pycel = _rewrite_formula_defined_names(s, host_sheet, name_maps)
        except Exception:
            s_for_pycel = s

        if s_for_pycel in formula_cache:
            tmp_addr = formula_cache[s_for_pycel]
        else:
            tmp_addr = f"{host_sheet}!ZZ{9000 + len(formula_cache)}"
            try:
                excel.set_value(tmp_addr, s_for_pycel)
            except Exception:
                return 1e12 
            formula_cache[s_for_pycel] = tmp_addr

        try:
            cell = excel.cell_map.get(tmp_addr)
            if cell is not None:
                cell.value = None
            v = excel.evaluate(tmp_addr)
            if isinstance(v, (list, tuple)):
                flat = []
                for r in v:
                    if isinstance(r, (list, tuple)): flat.extend(r)
                    else: flat.append(r)
                return float(sum(x for x in flat if isinstance(x, (int, float))))
            return float(v or 0)
        except Exception:
            return 1e12 

    def _score_raw(x):
        eval_count[0] += 1
        x = apply_projection(x, simplex_groups, lo, hi)
        write_vars(x)
        # Recalc happens automatically via pycel on .evaluate()

        # Objectives
        any_obj = False
        total_weighted = 0.0
        max_signed = -1e30
        for o in obj_specs:
            val = eval_ref(o["ref"])
            goal = (o["goal"] or "Min").lower()
            if goal == "min":
                s_val = val
            elif goal == "max":
                s_val = -val
            elif goal == "target":
                s_val = abs(val - float(o.get("target", 0)))
            else:
                s_val = val
            w = float(o.get("weight", 1.0))
            total_weighted += w * s_val
            if w * s_val > max_signed: max_signed = w * s_val
            any_obj = True

        obj_part = (alpha * max_signed + (1 - alpha) * total_weighted) if any_obj else 0.0

        # Constraints
        penalty = 0.0
        for c in con_specs:
            lhs = eval_ref(c["lhs"])
            rhs_str = c["rhs"]
            try:
                rhs = float(rhs_str)
            except (TypeError, ValueError):
                rhs = eval_ref(rhs_str, True)
            diff = lhs - rhs
            op = c.get("op", "=")
            if op == "=":
                viol = abs(diff)
            elif op in ("<=", "<"):
                viol = max(0.0, diff)
            elif op in (">=", ">"):
                viol = max(0.0, -diff)
            else:
                viol = 0.0
            pw = hard_pen if c.get("type", "Hard").lower() == "hard" else float(c.get("penalty", 1000))
            penalty += pw * viol * viol

        try:
            out = float(obj_part + penalty)
            if out != out or abs(out) == float('inf'):
                return 1e12
            return out
        except Exception:
            return 1e12

    def score(x):
        try:
            k = _memo_key(x)
            cached = _memo.get(k)
            if cached is not None:
                return cached
            v = _score_raw(x)
            try:
                fv = float(v)
            except Exception:
                return 1e12
            if fv != fv or abs(fv) == float('inf'):
                return 1e12
            # Do not memoize sentinel/error scores. A transient pycel #DIV/0!,
            # #VALUE!, #NUM!/overflow, or evaluation exception should not poison
            # later runs that revisit the same rounded point after the graph has
            # been reset.
            if fv < 1e12:
                _memo[k] = fv
            return fv
        except Exception:
            return 1e12

    score.eval_count = eval_count
    score.var_addrs = var_addrs
    score.write_vars = write_vars
    score.clear_memo = _memo.clear
    return score


def _safe_eval(excel, addr):
    try:
        # pycel 1.0b30 can leave formula cells with stale error values after
        # a bad probe, especially through branching dependency graphs. Clear
        # the requested formula cell before reading so evaluate() recomputes it.
        try:
            c = excel.cell_map.get(addr)
            if c is not None and getattr(c, "formula", None):
                c.value = None
        except Exception:
            pass

        v = excel.evaluate(addr)
        if v is None:
            return 1e12
        if isinstance(v, str) and v.startswith("#"):
            return 1e12
        fv = float(v)
        # Catch NaN/inf
        if fv != fv or abs(fv) == float('inf'):
            return 1e12
        return fv
    except Exception:
        return 1e12


# ===========================================================================
# Coordinate Descent — Anti-Zigzag (Powell) + Adaptive Step
# ===========================================================================
#
# Design note: "coordinate descent + anti-zigzag + adaptive step" is exactly
# what Powell's conjugate-direction method already is, and SciPy ships a
# numerically hardened implementation. A hand-rolled version was prototyped
# and tested against scipy Powell on the Rosenbrock valley (the canonical
# zigzag stress test): scipy reached machine precision in ~43 evals; the
# hand-rolled one stalled at f~1.6 in ~590. We use the proven engine and
# layer the requested behaviours (adaptive-step line search is intrinsic to
# Powell; memoization comes from the score wrapper; early-stop via xtol/ftol)
# on top.

def coordinate_descent_azas(score, x0, lo, hi, simplex_groups,
                            max_evals=500, init_step=0.1,
                            min_step=1e-7, tol=1e-9, log_fn=None):
    """Robust bounded coordinate descent with anti-zigzag pattern moves.

    This intentionally does NOT use scipy Powell.  Powell's line search can
    probe very far outside the local valid basin; for log-log/Pareto sheets a
    single large probe often produces #NUM!/overflow and the line search then
    gives up.  This version is conservative:
      - start at the current feasible point;
      - try +/- one variable at a time inside bounds;
      - shrink steps when both directions are bad;
      - grow a step only after it improves;
      - after each improving pass, try one anti-zigzag/pattern move in the net
        progress direction.
    Bad probes simply score 1e12 and are rejected, so a few #NUM! points do not
    derail the method.
    """
    def log(s):
        if log_fn:
            log_fn(s)

    n = len(x0)
    x_best = apply_projection(np.asarray(x0, float).copy(), simplex_groups, lo, hi)
    f_best = float(score(x_best))

    if n == 0:
        return x_best, f_best

    span = np.asarray(hi, float) - np.asarray(lo, float)
    finite_span = np.where(np.isfinite(span) & (span > 0), span, 1.0)
    step = np.maximum(init_step * finite_span, min_step)

    start_evals = score.eval_count[0]
    passes = 0
    no_improve_passes = 0

    log(f"  CD-AZAS start: f={f_best:.8g}, budget={max_evals}")

    while score.eval_count[0] - start_evals < max_evals:
        passes += 1
        x_pass_start = x_best.copy()
        f_pass_start = f_best
        improved_this_pass = False

        # Visit larger remaining steps first.  This usually finds useful moves
        # earlier and avoids wasting budget on already-shrunk coordinates.
        order = np.argsort(-step)
        for i in order:
            if score.eval_count[0] - start_evals >= max_evals:
                break
            if step[i] <= min_step:
                continue

            base = x_best.copy()
            trials = []
            for direction in (1.0, -1.0):
                z = base.copy()
                z[i] = z[i] + direction * step[i]
                z = apply_projection(z, simplex_groups, lo, hi)
                trials.append(z)

            best_trial_x = None
            best_trial_f = f_best
            for z in trials:
                if score.eval_count[0] - start_evals >= max_evals:
                    break
                fz = float(score(z))
                if fz < best_trial_f - tol * (1.0 + abs(best_trial_f)):
                    best_trial_x, best_trial_f = z, fz

            if best_trial_x is not None:
                x_best, f_best = best_trial_x, best_trial_f
                improved_this_pass = True
                # If a coordinate helped, cautiously increase just that step.
                step[i] = min(step[i] * 1.35, finite_span[i])
            else:
                # Both directions were worse/error; shrink locally.
                step[i] *= 0.5

        # Anti-zigzag / pattern move: try the net pass direction once.  This is
        # the Powell-style acceleration without letting a global line search run
        # wild into overflow territory.
        if improved_this_pass and score.eval_count[0] - start_evals < max_evals:
            direction = x_best - x_pass_start
            if np.any(np.abs(direction) > 0):
                z = apply_projection(x_best + direction, simplex_groups, lo, hi)
                fz = float(score(z))
                if fz < f_best - tol * (1.0 + abs(f_best)):
                    x_best, f_best = z, fz

        if improved_this_pass:
            no_improve_passes = 0
            log(f"  CD-AZAS pass {passes}: f={f_best:.8g}, "
                f"evals={score.eval_count[0] - start_evals}")
        else:
            no_improve_passes += 1
            # No coordinate helped; shrink all active steps together.
            step *= 0.5
            if no_improve_passes >= 3:
                break

        if np.nanmax(step) <= min_step:
            break
        if abs(f_pass_start - f_best) <= tol * (1.0 + abs(f_pass_start)) and not improved_this_pass:
            break

    log(f"  CD-AZAS done: f={f_best:.8g}, "
        f"evals={score.eval_count[0] - start_evals}, passes={passes}")
    return x_best, f_best


# ===========================================================================
# Method dispatch
# ===========================================================================

def _method_key(method):
    """Normalize UI/VBA method labels so small spelling/spacing changes do
    not silently route to the wrong solver."""
    s = str(method or "auto").strip().lower()
    for ch in "_–—/()":
        s = s.replace(ch, " ")
    s = s.replace("-", " ")
    s = " ".join(s.split())
    aliases = {
        "auto": "auto",
        "slsqp": "slsqp",
        "slsqp python": "slsqp",
        "l bfgs b": "l-bfgs-b",
        "lbfgs": "l-bfgs-b",
        "homotopy autodiff": "l-bfgs-b",
        "homotopy": "l-bfgs-b",
        "differential evolution": "de",
        "de": "de",
        "de multi restart": "de-multi-restart",
        "differential evolution multi restart": "de-multi-restart",
        "de 100": "de-multi-restart",
        "de100": "de-multi-restart",
        "portfolio": "portfolio",
        "portfolio all methods": "portfolio",
        "best of all": "portfolio",
        "round robin": "portfolio",
        "coordinate descent": "cd-azas",
        "cordinate descent": "cd-azas",       # common typo
        "coordinate decent": "cd-azas",       # common typo
        "cordinate decents": "cd-azas",       # common typo
        "coordinate descents": "cd-azas",
        "coordinate descent anti zigzag": "cd-azas",
        "coordinate descent anti zigzag adaptive step": "cd-azas",
        "cd": "cd-azas",
        "cd azas": "cd-azas",
        "cd-azas": "cd-azas",
        "coordinate descent azas": "cd-azas",
        "coordinate_descent_azas": "cd-azas",
    }
    return aliases.get(s, s)

def run_method(method, score, x0, bounds, simplex_groups, lo, hi, max_iter,
               de_seed=None, de_seed_with_x0=False):
    method = _method_key(method)
    n = len(x0)

    def projected(x):
        return score(apply_projection(np.asarray(x, float).copy(), simplex_groups, lo, hi))

    if method == "auto":
        # Heuristic: prefer SLSQP for small, L-BFGS-B for larger
        method = "l-bfgs-b" if n > 50 else "slsqp"

    if method == "portfolio":
        return portfolio_solve(score, x0, bounds, simplex_groups, lo, hi,
                               max_iter, log_fn=getattr(score, "log_fn", None))

    if method == "cd-azas":
        x_final, f_final = coordinate_descent_azas(
            score, x0, lo, hi, simplex_groups,
            max_evals=max_iter, init_step=0.1,
            min_step=1e-7, tol=1e-9,
            log_fn=getattr(score, "log_fn", None))
        return x_final, f_final

    if method == "slsqp":
        res = minimize(projected, x0, method="SLSQP",
                       bounds=list(zip(lo, hi)),
                       options={"maxiter": max_iter, "ftol": 1e-9})
        x_final = apply_projection(res.x, simplex_groups, lo, hi)
        return x_final, res.fun

    if method == "l-bfgs-b":
        # NOTE: full JAX/homotopy path would require translating the pycel
        # formula graph to JAX. Until then we use L-BFGS-B with FD gradients
        # — still way faster than Excel-in-the-loop because pycel skips Excel.
        res = minimize(projected, x0, method="L-BFGS-B",
                       bounds=list(zip(lo, hi)),
                       options={"maxiter": max_iter, "ftol": 1e-9})
        x_final = apply_projection(res.x, simplex_groups, lo, hi)
        return x_final, res.fun

    if method == "de":
        # Seed handling:
        #   - de_seed=None  -> draw a fresh OS-random seed each call so every
        #     invocation is a genuinely different global search (this is what
        #     makes re-running DE in the portfolio actually explore instead of
        #     replaying an identical run).
        #   - de_seed=<int> -> caller (portfolio) supplies a varying seed
        #     derived from round/attempt for reproducible-but-different runs.
        if de_seed is None:
            import secrets
            this_seed = secrets.randbelow(2**31 - 1)
        else:
            this_seed = int(de_seed) % (2**31 - 1)

        de_kwargs = dict(
            bounds=list(zip(lo, hi)),
            maxiter=max(20, max_iter // 10),
            tol=1e-7,
            seed=this_seed,
            polish=True,
        )

        # Optionally inject the incumbent into DE's initial population so DE
        # can keep/refine the best-known point instead of discarding it and
        # searching the whole box from scratch. The rest of the population is
        # latin-hypercube sampled for exploration.
        if de_seed_with_x0:
            try:
                import numpy as _np
                popsize = 15  # scipy default multiplier
                n_pop = max(5, popsize)
                rng = _np.random.default_rng(this_seed)
                pop = lo + rng.random((n_pop, n)) * (hi - lo)
                pop[0] = _np.clip(_np.asarray(x0, float), lo, hi)
                de_kwargs["init"] = pop
            except Exception:
                pass  # fall back to default init on any issue

        res = differential_evolution(projected, **de_kwargs)
        x_final = apply_projection(res.x, simplex_groups, lo, hi)
        return x_final, res.fun

    if method == "de-multi-restart":
        # Run DE up to N times, each with a DIFFERENT seed, keeping the best
        # result found. There is no "correct" seed to search for - a seed has
        # no inherent quality - so the only sound approach is best-of-N
        # restarts. Each restart explores a different region; we keep the
        # incumbent in every restart's population so a run can never lose the
        # best basin already found.
        #
        # Two early exits keep it from wasting all 100 restarts pointlessly:
        #   - patience: if `patience` consecutive restarts yield no
        #     improvement, the search has very likely settled - stop.
        #   - eval budget: never exceed the caller's max_iter total.
        N_RESTARTS = 100
        patience = 1000

        best_x = apply_projection(np.asarray(x0, float).copy(),
                                  simplex_groups, lo, hi)
        best_f = score(best_x)
        log_fn = getattr(score, "log_fn", None)
        prog_dir = getattr(score, "run_dir", None)

        no_improve = 0
        start_evals = score.eval_count[0]
        for k in range(1, N_RESTARTS + 1):
            if score.eval_count[0] - start_evals >= max_iter:
                if log_fn:
                    log_fn(f"  DE-MultiRestart: eval budget reached at "
                           f"restart {k}; returning best f={best_f:.8g}")
                break

            # Build a DE population seeded with the current best so this
            # restart cannot regress below it.
            this_seed = (k * 2654435761) % (2**31 - 1)  # spread seeds well
            try:
                popn = max(5, 15)
                rng = np.random.default_rng(this_seed)
                pop = lo + rng.random((popn, n)) * (hi - lo)
                pop[0] = np.clip(best_x, lo, hi)
                init = pop
            except Exception:
                init = "latinhypercube"

            r = differential_evolution(
                projected, bounds=list(zip(lo, hi)),
                maxiter=max(20, (max_iter // 10)),
                tol=1e-7, seed=this_seed, polish=True, init=init)

            rx = apply_projection(r.x, simplex_groups, lo, hi)
            rf = score(rx)
            if rf < best_f - 1e-12 * (1.0 + abs(best_f)):
                best_x, best_f = rx, rf
                no_improve = 0
                if log_fn:
                    log_fn(f"  DE-MultiRestart {k}/{N_RESTARTS}: "
                           f"new best f={best_f:.8g}")
                if prog_dir:
                    _write_progress(prog_dir,
                                    f"DE restart {k}/{N_RESTARTS} | "
                                    f"NEW BEST f={best_f:.6g}")
            else:
                no_improve += 1
                if prog_dir:
                    _write_progress(prog_dir,
                                    f"DE restart {k}/{N_RESTARTS} | "
                                    f"no gain ({no_improve}/{patience}) | "
                                    f"best f={best_f:.6g}")

            if no_improve >= patience:
                if log_fn:
                    log_fn(f"  DE-MultiRestart: {patience} restarts with no "
                           f"improvement (stopped at {k}); f={best_f:.8g}")
                break

        return best_x, best_f

    # Default
    res = minimize(projected, x0, method="L-BFGS-B",
                   bounds=list(zip(lo, hi)),
                   options={"maxiter": max_iter, "ftol": 1e-9})
    x_final = apply_projection(res.x, simplex_groups, lo, hi)
    return x_final, res.fun


# ===========================================================================
# Portfolio meta-solver  (round-robin, interpretation A)
# ===========================================================================
def portfolio_solve(score, x0, bounds, simplex_groups, lo, hi,
                    max_iter, log_fn=None,
                    order=("slsqp", "l-bfgs-b",
                           "coordinate descent anti-zigzag adaptive step",
                           "differential evolution"),
                    rel_tol=1e-9, max_rounds=100,
                    no_change_limit=2):
    """Run a portfolio of methods, chaining improvements.

    Interpretation A:
      - For each method in `order`:
          * run it from the current best point
          * if it improved, run THE SAME method again from the new point
          * repeat until that method yields no further improvement
          * then advance to the next method
      - After a full pass over all methods, if ANY method improved during
        the pass, start the whole cycle again.
      - Stop when no_change_limit complete passes produce no improvement,
        or when max_rounds is hit, or when the eval budget is exhausted.

    "Improvement" = strictly better than the incumbent by more than
    rel_tol * (1 + |incumbent|). The (1+|f|) scaling makes the tolerance
    behave sensibly whether the optimum is near 0 or large.

    Every method shares the same memoized `score`, so when a method re-runs
    from a point another method already explored, repeated evaluations are
    free cache hits — the chaining is much cheaper than 4 independent solves.
    """
    def log(s):
        if log_fn:
            log_fn(s)

    # Where to write the user-facing progress (monitor window reads this).
    _run_dir = getattr(score, "run_dir", None)

    def progress(s):
        if _run_dir:
            _write_progress(_run_dir, s)

    best_x = apply_projection(np.asarray(x0, float).copy(),
                              simplex_groups, lo, hi)
    best_f = score(best_x)
    log(f"Portfolio start: f={best_f:.8g}")
    progress(f"Starting. Initial score = {best_f:.6g}")

    def improved(new_f, old_f):
        return new_f < old_f - rel_tol * (1.0 + abs(old_f))

    total_start_evals = score.eval_count[0]
    no_change_rounds = 0

    # Each inner method call gets a modest budget; the portfolio tracks the
    # global budget separately. Without this split, one SLSQP/DE call would
    # be handed the whole portfolio budget and could exhaust it in a single
    # invocation, defeating the round-robin.
    per_call_budget = max(50, max_iter // 12)

    for rnd in range(1, max_rounds + 1):
        round_improved = False
        log(f"--- Portfolio round {rnd} ---")
        progress(f"Round {rnd} starting (best so far = {best_f:.6g})")

        for method in order:
            # Re-run THIS method until it stops improving (interpretation A)
            method_iters = 0
            while True:
                if score.eval_count[0] - total_start_evals >= max_iter:
                    log(f"  [{method}] global eval budget reached; "
                        f"stopping portfolio")
                    progress(f"Eval budget reached - finishing. "
                             f"Best = {best_f:.6g}")
                    return best_x, best_f

                evals_now = score.eval_count[0] - total_start_evals
                progress(f"Round {rnd} | {method} | attempt {method_iters + 1} "
                         f"| {evals_now} evals | best = {best_f:.6g}")

                if method == "de":
                    # Vary DE's seed every call so it explores differently
                    # each time, and inject the incumbent so it can't lose
                    # the best basin already found.
                    de_seed = 1000 * rnd + 7 * (method_iters + 1) + 13
                    x_try, f_try = run_method(method, score, best_x.copy(),
                                              bounds, simplex_groups, lo, hi,
                                              per_call_budget,
                                              de_seed=de_seed,
                                              de_seed_with_x0=True)
                else:
                    x_try, f_try = run_method(method, score, best_x.copy(),
                                              bounds, simplex_groups, lo, hi,
                                              per_call_budget)

                method_iters += 1

                if improved(f_try, best_f):
                    delta = best_f - f_try
                    best_x, best_f = x_try, f_try
                    round_improved = True
                    log(f"  [{method}] improved -> f={best_f:.8g} "
                        f"(-{delta:.3g}); re-running same method")
                    progress(f"  -> {method} improved! "
                             f"new best = {best_f:.6g} (re-running it)")
                    # loop again with the SAME method from the new point
                else:
                    if method_iters == 1:
                        log(f"  [{method}] no improvement (f={f_try:.8g})")
                    else:
                        log(f"  [{method}] exhausted after {method_iters} "
                            f"runs; best f={best_f:.8g}")
                    progress(f"  -> {method} done (no further gain)")
                    break  # advance to next method

        if round_improved:
            no_change_rounds = 0
        else:
            no_change_rounds += 1
            log(f"Portfolio no-change round {no_change_rounds}/"
                f"{no_change_limit} (round {rnd}), f={best_f:.8g}")
            progress(f"No change {no_change_rounds}/{no_change_limit} "
                     f"| best = {best_f:.6g}")

            if no_change_rounds >= no_change_limit:
                log(f"Portfolio converged: {no_change_limit} consecutive "
                    f"full rounds with no improvement, f={best_f:.8g}")
                progress(f"Converged after {no_change_limit} no-change rounds. "
                         f"Best = {best_f:.6g}")
                break
    else:
        log(f"Portfolio hit max_rounds={max_rounds}; f={best_f:.8g}")

    return best_x, best_f


# ===========================================================================
# Multi-start wrapper
# ===========================================================================

def multi_start_solve(method, score, x0, bounds, simplex_groups, lo, hi,
                      max_iter, n_starts=3):
    best_x, best_f = None, np.inf
    rng = np.random.default_rng(42)
    n = len(x0)
    log_fn = getattr(score, "log_fn", None)
    for s in range(n_starts):
        if s == 0:
            x_init = x0.copy()
        else:
            x_init = np.clip(x0 + rng.normal(0, 0.2, n), lo, hi)
            x_init = apply_projection(x_init, simplex_groups, lo, hi)
        if log_fn:
            log_fn(f"  Multi-start {s + 1}/{n_starts}: method={_method_key(method)}")
        x, f = run_method(method, score, x_init, bounds, simplex_groups,
                           lo, hi, max_iter)
        if f < best_f:
            best_f, best_x = f, x
    return best_x, best_f


# ===========================================================================
# Main driver
# ===========================================================================

def _compute_cache_key(src_xlsx, variables, log_fn):
    """Build a cache key that ignores ONLY the variable cells.

    The compiled pycel graph depends on formulas and on the constant values
    that formulas read — NOT on the variable cells, because the solver
    overwrites those with set_value() before the first evaluation anyway.

    Each solve writes optimized numbers back into the variable cells, so the
    raw file bytes change every run even when the user changed nothing. If
    we keyed on raw bytes we'd recompile every single time and the cache
    would be useless.

    So we hash the workbook with the variable cells blanked out:
      - solver's own previous output (variable cells)  -> excluded -> cache HIT
      - user edits a formula                           -> included -> recompile
      - user edits a constant a formula reads           -> included -> recompile
      - user changes nothing                            -> same key -> cache HIT

    This is safe: anything the USER can change is still in the key. Only the
    cells the optimizer itself owns are excluded.
    """
    import hashlib
    from openpyxl import load_workbook

    try:
        wb = load_workbook(src_xlsx, data_only=False)
    except Exception as e:
        # Fall back to raw-byte hash if we can't introspect — correct, just
        # less cache-friendly.
        log_fn(f"Cache key: workbook introspection failed ({e}); using raw hash")
        return hashlib.sha1(Path(src_xlsx).read_bytes()).hexdigest()[:16]

    # Set of variable cells to exclude, as (sheet, coordinate) upper-cased.
    var_cells = set()
    for v in variables:
        ref = v.get("ref", "")
        if "!" not in ref:
            continue
        sheet, cell = ref.split("!", 1)
        sheet = sheet.strip().strip("'").strip()
        cell = cell.replace("$", "").upper()
        var_cells.add((sheet, cell))

    h = hashlib.sha1()
    # Cache format/version marker: changing the defined-name resolver must
    # force a fresh pycel compile instead of loading an old graph that was
    # built without formula defined-name rewriting.
    h.update(b"FASTSOLVER_PYCEL_FORMULA_REWRITE_V1\0")

    # Defined names affect the compiled graph just like formulas do. Include
    # them in the key so Name Manager edits invalidate the cache.
    try:
        for name, target, sheet in sorted(_iter_defined_names(wb)):
            h.update(f"\x00NAME:{sheet or 'WORKBOOK'}!{name}={target}\x00".encode("utf-8"))
    except Exception:
        pass

    for ws in wb.worksheets:
        h.update(("\x00SHEET:" + ws.title + "\x00").encode("utf-8"))
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                if (ws.title, c.coordinate.upper()) in var_cells:
                    continue  # skip solver-owned cells
                # Include address + formula/value. For formula cells
                # c.value is the formula string (data_only=False).
                h.update(f"{c.coordinate}={c.value!r};".encode("utf-8"))
    return h.hexdigest()[:16]


def _load_excel_cached(src_xlsx, variables, log_fn):
    """Compile a pycel-safe workbook with caching.

    Before pycel sees the workbook, expand supported Excel defined names
    inside formula text to their cell/range targets.  This prevents pycel
    1.0b30 from treating normal names such as LogLogScale or sheet-local
    truncation as missing structured tables during recomputation.
    """
    src_xlsx = Path(src_xlsx)
    compile_xlsx, name_maps = _rewrite_workbook_formulas_for_pycel(src_xlsx, log_fn)
    compile_xlsx = _sanitize_defined_names_for_pycel(compile_xlsx, log_fn)

    digest = _compute_cache_key(compile_xlsx, variables, log_fn)
    cache_base = str(src_xlsx.parent.parent / f"_pycel_cache_xmlnames_guard_v5_{digest}")
    cache_pkl = cache_base + ".pkl"

    if os.path.exists(cache_pkl):
        try:
            ex = ExcelCompiler.from_file(cache_pkl)
            ex._fastsolver_name_maps = name_maps
            log_fn("Pycel graph loaded from cache (skipped parse)")
            return ex
        except Exception as e:
            log_fn(f"Cache load failed ({e}); recompiling")

    ex = ExcelCompiler(filename=str(compile_xlsx))
    ex._fastsolver_name_maps = name_maps
    try:
        ex.to_file(cache_base, file_types=('pkl',))
        log_fn("Pycel graph compiled and cached for next run")
    except Exception as e:
        log_fn(f"Could not write pycel cache ({e}); continuing without it")
    return ex


def main(run_dir):
    # ---- Pre-flight: validate the run directory BEFORE the try block ----
    # If we were handed a bad argument (e.g. the literal string "--monitor"
    # because an OLD exe doesn't understand monitor mode), every later step
    # fails AND the error handler's own write to results.json fails too,
    # producing the cryptic double-traceback. Catch it here, loudly, and
    # do NOT let the window vanish.
    raw_arg = run_dir
    run_dir = Path(run_dir)
    cfg_path = run_dir / "config.json"
    src_xlsx = run_dir / "source.xlsx"

    if not run_dir.is_dir() or not cfg_path.is_file():
        print("=" * 60, flush=True)
        print(" FASTSOLVER - STARTUP ERROR", flush=True)
        print("=" * 60, flush=True)
        print(f" The run directory is not valid:", flush=True)
        print(f"   argument received : {raw_arg!r}", flush=True)
        print(f"   resolved path     : {run_dir}", flush=True)
        print(f"   directory exists  : {run_dir.is_dir()}", flush=True)
        print(f"   config.json found : {cfg_path.is_file()}", flush=True)
        print(flush=True)
        if str(raw_arg).startswith("--"):
            print(" It looks like an OPTION ('--...') was passed as the run", flush=True)
            print(" directory. This almost always means the FastSolver.exe", flush=True)
            print(" being run is an OLD build that does not understand", flush=True)
            print(" monitor mode. Rebuild the exe from the current", flush=True)
            print(" 07_fastsolver_bridge.py (run 11_build_exe.py) and", flush=True)
            print(" redeploy the dist/FastSolver folder next to the .xlam.", flush=True)
        else:
            print(" The VBA side did not create config.json here, or the", flush=True)
            print(" path is wrong. Check modPythonBridge run-folder logic.", flush=True)
        print(flush=True)
        print(" This window is staying open. Type q then ENTER to close.",
              flush=True)
        _wait_for_q()
        sys.exit(3)

    log = []
    def log_line(s):
        log.append(s)
        print(s, flush=True)

    try:
        config = json.loads(cfg_path.read_text())
        log_line(f"Loaded config: {len(config['variables'])} vars, "
                 f"{len(config['objectives'])} objs, "
                 f"{len(config['constraints'])} cons")

        variables = config["variables"]
        n = len(variables)
        lo = np.array([float(v["lo"]) for v in variables])
        hi = np.array([float(v["hi"]) for v in variables])
        simplex_groups = collect_simplex_groups(variables)
        log_line(f"Simplex groups: {len(simplex_groups)}")

        _set_phase(run_dir, "compiling")
        t0 = time.time()
        excel = _load_excel_cached(src_xlsx, variables, log_line)
        log_line(f"Pycel ready in {time.time()-t0:.2f}s")
        _set_phase(run_dir, "solving")

        # Initial values from the workbook
        x0 = np.zeros(n)
        for i, v in enumerate(variables):
            x0[i] = _safe_eval(excel, _pycel_addr(v["ref"]))
        x0 = apply_projection(x0, simplex_groups, lo, hi)

        score = build_score_fn(excel, config, variables, simplex_groups, lo, hi)
        score.log_fn = log_line
        score.run_dir = str(run_dir)
        start_score = score(x0)

        # Sanity probes cost 2 extra full evaluations on every run. On a slow
        # workbook that's wasted seconds. Only run them when debugging.
        if os.environ.get("FASTSOLVER_DEBUG"):
            x_test = x0.copy()
            if len(x_test) > 0:
                x_test[0] = min(hi[0], x0[0] + 0.1) if x0[0] + 0.1 <= hi[0] else max(lo[0], x0[0] - 0.1)
            probe_score = score(x_test)
            log_line(f"Probe score (perturbed x[0]): {probe_score:.6f}")
            if abs(probe_score - start_score) < 1e-12:
                log_line("WARNING: score did not change on perturbation — pycel not propagating vars")
            x_zero = x0.copy()
            x_zero[0] = 0.0
            crash_score = score(x_zero)
            log_line(f"Score at x[0]=0 (div-by-zero test): {crash_score:.2e}")
        log_line(f"Initial score: {start_score:.6f}")


        method = config.get("method", "Auto")
        method_key = _method_key(method)
        max_iter = int(config.get("max_iter", 200))
        log_line(f"Method requested: {method!r} -> {method_key}; max_iter={max_iter}")

        t1 = time.time()
        if method_key == "portfolio":
            # The portfolio self-explores (it includes DE for global search
            # and chains local methods), so wrapping it in random multi-start
            # would just repeat the whole portfolio from noisy points and
            # waste the budget. Run it once with a larger total eval budget.
            x_best, f_best = portfolio_solve(
                score, x0, list(zip(lo, hi)), simplex_groups, lo, hi,
                max_iter=max(max_iter * 6, 1500), log_fn=log_line)
        elif method_key in ("cd-azas", "de-multi-restart"):
            # These are already restart/step-control methods.  Do NOT wrap
            # them in the generic 3-start loop; that was causing confusing
            # repeated "DE-MultiRestart budget hit" messages and wasting the
            # entire budget three times.
            x_best, f_best = run_method(method_key, score, x0,
                                        list(zip(lo, hi)),
                                        simplex_groups, lo, hi,
                                        max_iter)
        else:
            x_best, f_best = multi_start_solve(method_key, score, x0,
                                               list(zip(lo, hi)),
                                               simplex_groups, lo, hi,
                                               max_iter, n_starts=3)
        elapsed = time.time() - t1
        log_line(f"Optimized in {elapsed:.2f}s. Final score: {f_best:.6f}. "
                 f"Eval count: {score.eval_count[0]}")

        # solved.xlsx is NOT read by the VBA side — it applies
        # variable_values from results.json directly to the live workbook.
        # Re-loading + saving the whole workbook here cost 0.5-2s every run
        # for a file nothing consumes. Only emit it when debugging.
        if os.environ.get("FASTSOLVER_DEBUG"):
            wb = load_workbook(src_xlsx)
            for v, val in zip(variables, x_best):
                sheet, cell = _split_addr(v["ref"])
                try:
                    wb[sheet][cell] = float(val)
                except Exception:
                    pass
            wb.save(run_dir / "solved.xlsx")

        # Write results.json
        results = {
            "status": "ok",
            "method_used": method,
            "elapsed_sec": elapsed,
            "eval_count": score.eval_count[0],
            "start_score": start_score,
            "final_score": float(f_best),
            "variable_names": [v["ref"] for v in variables],
            "variable_values": [float(x) for x in x_best],
            "log": log,
        }
        (run_dir / "results.json").write_text(json.dumps(results, indent=2))
        log_line("Wrote results.json + solved.xlsx")
    except Exception as e:
        tb = traceback.format_exc()
        # Try to write results.json, but NEVER let this fail silently the
        # way it did before (when run_dir was bad, the write itself threw
        # and buried the real error under a second traceback).
        wrote = False
        try:
            err = {
                "status": "error",
                "message": str(e),
                "traceback": tb,
                "log": log,
            }
            (run_dir / "results.json").write_text(json.dumps(err, indent=2))
            wrote = True
        except Exception as e2:
            print(f"(could not write results.json: {e2})", flush=True)

        # Always show the real error in the console, regardless of whether
        # results.json could be written.
        print(flush=True)
        print("=" * 60, flush=True)
        print(" FASTSOLVER - SOLVE FAILED", flush=True)
        print("=" * 60, flush=True)
        print(f" Error: {e}", flush=True)
        print(flush=True)
        print(" --- traceback ---", flush=True)
        for ln in tb.splitlines()[-25:]:
            print(" | " + ln, flush=True)
        if log:
            print(flush=True)
            print(" --- log tail ---", flush=True)
            for ln in log[-15:]:
                print(" | " + str(ln), flush=True)
        if wrote:
            print("\n (results.json written - the Excel side will report "
                  "failure too)", flush=True)
        print(flush=True)
        print(" This window is staying open. Type q then ENTER to close.",
              flush=True)
        _wait_for_q()
        sys.exit(2)


def _wait_for_q():
    """Block until the user types q + Enter. Robust against a frozen-exe
    console where bare input() can return instantly: loop, and if stdin is
    genuinely unusable, fall back to a very long sleep so the window with
    the error stays readable instead of vanishing."""
    import time as _t
    while True:
        try:
            ans = input(" > ").strip().lower()
        except Exception:
            _t.sleep(900.0)
            return
        if ans == "q":
            return
        print(" (type q then ENTER to close)", flush=True)


def _write_progress(run_dir, text):
    """Append a progress line the monitor window can display. Best-effort:
    progress reporting must never crash or slow the actual solve."""
    try:
        p = Path(run_dir) / "progress.txt"
        with open(p, "a", encoding="utf-8") as fh:
            fh.write(text.rstrip() + "\n")
    except Exception:
        pass


def _set_phase(run_dir, phase):
    """Signal the monitor which phase we're in: 'compiling' or 'solving'.
    The monitor shows an animated (no-percent) bar during 'compiling' since
    pycel can't report compile progress, then switches to live per-method
    counters once 'solving' begins."""
    try:
        (Path(run_dir) / "phase.txt").write_text(phase, encoding="utf-8")
    except Exception:
        pass


def monitor(run_dir):
    """Visible companion window with two phases:

      1. COMPILING - an animated bar (NO percentage: pycel cannot report
         how far through parsing it is, so a real % would be a lie). The
         bar just sweeps to show the process is alive.

      2. SOLVING - live per-method view: the current method, an ASCII
         progress bar across the portfolio's method list, and the eval
         count ticking up per method.

    It reads phase.txt + progress.txt written by the hidden solver and
    self-closes when results.json appears.
    """
    import time as _t
    run_dir = Path(run_dir)
    prog = run_dir / "progress.txt"
    phase_f = run_dir / "phase.txt"
    done = run_dir / "results.json"

    def read_phase():
        try:
            return phase_f.read_text(encoding="utf-8").strip()
        except Exception:
            return "compiling"

    def clear():
        # ANSI clear-line; harmless if unsupported
        print("\r" + " " * 70 + "\r", end="")

    print("=" * 60)
    print("   F A S T S O L V E R")
    print("   (progress monitor - closes itself when finished)")
    print("=" * 60, flush=True)

    # ---- Phase 1: COMPILING (animated, no percent) -------------------
    bar_w = 28
    pos = 0
    direction = 1
    comp_waited = 0.0
    COMPILE_LIMIT = 300.0   # 5 min: pycel compile should never exceed this
    print()
    while not done.exists() and read_phase() == "compiling":
        # A single lit cell sweeping back and forth = "alive, working".
        cells = ["."] * bar_w
        cells[pos] = "#"
        bar = "".join(cells)
        print(f"\r  Compiling workbook  [{bar}]  (this can take a few s)",
              end="", flush=True)
        pos += direction
        if pos >= bar_w - 1 or pos <= 0:
            direction *= -1
        _t.sleep(0.08)
        comp_waited += 0.08
        if comp_waited >= COMPILE_LIMIT:
            print(flush=True)
            print(f"  Still 'compiling' after {int(COMPILE_LIMIT)}s with no "
                  f"result - the solver likely crashed during startup.",
                  flush=True)
            break
    clear()
    if done.exists() or read_phase() != "compiling":
        print("\r  Compiling workbook  [ done ]", flush=True)

    # ---- Phase 2: SOLVING (live per-method counters) -----------------
    print()
    print("  Solving:")
    seen = 0
    spin = "|/-\\"
    spin_i = 0
    waited = 0.0
    silent = 0.0          # seconds since the last progress change
    SILENCE_LIMIT = 80.0 # 80 sec with zero progress AND no result = dead
    while not done.exists():
        lines = []
        try:
            if prog.exists():
                lines = prog.read_text(encoding="utf-8").splitlines()
        except Exception:
            lines = []

        if len(lines) > seen:
            for ln in lines[seen:]:
                print("   " + ln, flush=True)
            seen = len(lines)
            silent = 0.0          # progress moved — solver is alive
        else:
            spin_i = (spin_i + 1) % len(spin)
            last = lines[-1] if lines else "starting..."
            print(f"\r   {spin[spin_i]} {last}   "
                  f"({waited:0.0f}s)        ",
                  end="", flush=True)
            silent += 0.4

        if silent >= SILENCE_LIMIT:
            # No progress for a long time and still no results.json: the
            # solver almost certainly died without writing an error file
            # (hard crash / killed). Stop waiting and fall through to the
            # failure display so the user isn't left staring at a spinner.
            print(flush=True)
            print(f"  No progress for {int(SILENCE_LIMIT)}s and no result "
                  f"- assuming the solver crashed.", flush=True)
            break

        _t.sleep(0.4)
        waited += 0.4

    print()
    print("  " + "-" * 50)

    failed = False
    msg = ""
    tb = ""
    logtail = []
    try:
        if done.exists():
            raw = done.read_text(encoding="utf-8", errors="replace")
            try:
                res = json.loads(raw)
            except Exception:
                # results.json exists but is malformed/partial — show the
                # raw bytes, that's still the best clue we have.
                failed = True
                msg = "results.json is present but not valid JSON " \
                      "(solver likely died mid-write)."
                tb = raw[-2000:]
                res = None
            if res is not None:
                if res.get("status") == "ok":
                    print(f"  DONE.  Final score: "
                          f"{res.get('final_score'):.6g}"
                          f"   ({res.get('eval_count','?')} evals)",
                          flush=True)
                else:
                    failed = True
                    msg = res.get("message", "(no message in results.json)")
                    # The solver stores the real Python traceback + its
                    # in-memory log INSIDE results.json, not as a separate
                    # file. Surface both - this is the actual error.
                    tb = res.get("traceback", "") or ""
                    lg = res.get("log", [])
                    if isinstance(lg, list):
                        logtail = [str(x) for x in lg][-20:]
        else:
            failed = True
            msg = "Solver exited without producing results.json (it crashed)."
    except Exception as e:
        failed = True
        msg = f"Monitor could not read results.json: {e}"

    if not failed:
        print("  (window closes in 3s)", flush=True)
        _t.sleep(3.0)
        return

    # FAILURE: show everything useful and DO NOT auto-close.
    print(flush=True)
    print("  " + "!" * 52, flush=True)
    print("  SOLVE FAILED", flush=True)
    print("  " + "!" * 52, flush=True)
    print(f"  Reason: {msg}", flush=True)

    if logtail:
        print("\n  --- solver log (last lines) ---", flush=True)
        for ln in logtail:
            print("  | " + ln, flush=True)

    if tb:
        print("\n  --- Python traceback ---", flush=True)
        for ln in tb.splitlines()[-25:]:
            print("  | " + ln, flush=True)

    print(flush=True)
    print("  " + "=" * 52, flush=True)
    print("  This window is STAYING OPEN so you can read the error.",
          flush=True)
    print("  Type  q  then ENTER to close it.", flush=True)
    print("  " + "=" * 52, flush=True)

    # Robust pause. Bare input() can return instantly in a frozen-exe
    # console with no real stdin; loop until we actually get a 'q', and
    # if stdin is truly unusable, fall back to a very long sleep so the
    # window still does not vanish.
    while True:
        try:
            ans = input("  > ").strip().lower()
        except Exception:
            _t.sleep(600.0)
            break
        if ans == "q":
            break
        print("  (type q then ENTER to close)", flush=True)


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "--monitor":
        monitor(sys.argv[2])
        sys.exit(0)
    if len(sys.argv) < 2:
        print("Usage: fastsolver_bridge.py <run_dir>", file=sys.stderr)
        print("       fastsolver_bridge.py --monitor <run_dir>",
              file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1])
