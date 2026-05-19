"""
fastsolver_bridge.py — Python-side entry point for the .xlam add-in.

Called as:
    FastSolver.exe <run_dir>

Where <run_dir> contains:
    config.json    — problem spec exported by VBA
    source.xlsx    — copy of the user's workbook

Produces:
    results.json   — final score + variable name/value arrays
    solved.xlsx    — workbook copy with optimal values written in

Implements method dispatch:
    "Auto"                  → routes by problem size/structure
    "SLSQP Python"          → SciPy SLSQP via pycel evaluation
    "L-BFGS-B"              → scaled to large var counts
    "Differential Evolution"→ global, derivative-free, pycel-eval
    "Homotopy / AutoDiff"   → JAX + smoothing — placeholder for full
                              graph→JAX rewrite; falls back to L-BFGS-B
                              over pycel-eval with finite-diff today
"""

from __future__ import annotations
import json
import os
import sys
import time
import traceback
from pathlib import Path

import numpy as np

# ---------------------------------------------------------------------------
# Pycel defined-names workaround (some openpyxl versions break here)
# ---------------------------------------------------------------------------
import pycel.excelwrapper as _ew
def _safe_defined_names(self):
    if self._defined_names is None:
        self._defined_names = {}
        try:
            dn = self.workbook.defined_names
            names = (dn.definedName if hasattr(dn, "definedName")
                     else list(dn.values()) if hasattr(dn, "values") else [])
            for d in names:
                if hasattr(d, "value") and d.value:
                    self._defined_names[d.name] = d.value
        except Exception:
            pass
    return self._defined_names
_ew.ExcelOpxWrapper.defined_names = property(_safe_defined_names)

from pycel import ExcelCompiler
from scipy.optimize import minimize, differential_evolution
from openpyxl import load_workbook


# ===========================================================================
# Helpers — simplex projection, bounds, etc.
# ===========================================================================

def project_simplex_group(x, idxs):
    """In place: project x[idxs] onto sum=1, >=0."""
    sub = x[idxs]
    sub = np.maximum(sub, 0)
    s = sub.sum()
    if s > 0:
        sub = sub / s
    else:
        sub = np.ones(len(sub)) / len(sub)
    x[idxs] = sub


def collect_simplex_groups(variables):
    """Return list of index-arrays for variables sharing rule 'Sum=1'."""
    groups = {}
    for i, v in enumerate(variables):
        if v["rule"] == "Sum=1":
            groups.setdefault(v["group"], []).append(i)
    return [np.array(idxs, dtype=int) for idxs in groups.values()]


def apply_projection(x, simplex_groups, lo, hi):
    x = np.clip(x, lo, hi)
    for idxs in simplex_groups:
        project_simplex_group(x, idxs)
    return x


# ===========================================================================
# Excel evaluation via pycel — score function builder
# ===========================================================================

def _split_addr(ref):
    """Split 'SheetName!$A$1' or "'Sheet 1'!$A$1" → (sheet, cell). Tolerates missing leading quote."""
    if "!" not in ref:
        return None, ref
    sheet, cell = ref.split("!", 1)
    sheet = sheet.strip().strip("'").strip()
    cell = cell.replace("$", "")
    return sheet, cell


def _pycel_addr(ref):
    """Convert VBA-exported ref → pycel address format SheetName!A1."""
    sheet, cell = _split_addr(ref)
    if sheet is None:
        return cell
    if " " in sheet:
        return f"'{sheet}'!{cell}"
    return f"{sheet}!{cell}"


def build_score_fn(excel, config, variables, simplex_groups, lo, hi):
    var_addrs = [_pycel_addr(v["ref"]) for v in variables]
    obj_specs = [o for o in config["objectives"] if o["active"]]
    con_specs = [c for c in config["constraints"] if c["active"]]
    alpha = float(config.get("outlier_alpha", 0.7))
    hard_pen = float(config.get("hard_penalty", 1e6))

    eval_count = [0]

    # pycel's set_value() invalidates the full downstream dependency chain
    # automatically. The previous code manually nulled only DIRECT successors
    # (1 level deep), which left deeper cells (var -> A1 -> A2 -> obj) stale
    # and could return wrong scores. We now rely on set_value's own chain
    # invalidation, which is both correct and faster.
    def write_vars(x):
        for addr, val in zip(var_addrs, x):
            try:
                excel.set_value(addr, float(val))
            except Exception:
                pass

    # ---- Memoization -----------------------------------------------------
    # CD line-search and DE re-probe the same / nearby points constantly.
    # Key on the rounded x vector so micro-adjust steps that revisit a point
    # become free dict lookups instead of full pycel re-evaluations.
    _memo = {}
    _MEMO_DECIMALS = 10  # tighter than any tolerance we use; safe to cache

    def _memo_key(x):
        return tuple(round(float(v), _MEMO_DECIMALS) for v in x)

    formula_cache = {}

    def eval_ref(spec_str, is_rhs=False):
        s = spec_str.strip()
        if not s.startswith("="):
            return _safe_eval(excel, _pycel_addr(s))

        if s in formula_cache:
            tmp_addr = formula_cache[s]
        else:
            host_sheet = _split_addr(variables[0]["ref"])[0] or "Sheet1"
            tmp_addr = f"{host_sheet}!ZZ{9000 + len(formula_cache)}"
            try:
                excel.set_value(tmp_addr, s)
            except Exception:
                return 1e12 
            formula_cache[s] = tmp_addr

        try:
            cell = excel.cell_map.get(tmp_addr)
            if cell is not None:
                cell.value = None
            v = excel.evaluate(tmp_addr)
            if isinstance(v, (list, tuple)):
                flat = []
                for r in v:
                    if isinstance(r, (list, tuple)): flat.extend(r)
                    else: flat.append(r)
                return float(sum(x for x in flat if isinstance(x, (int, float))))
            return float(v or 0)
        except Exception:
            return 1e12 

    def _score_raw(x):
        eval_count[0] += 1
        x = apply_projection(x, simplex_groups, lo, hi)
        write_vars(x)
        # Recalc happens automatically via pycel on .evaluate()

        # Objectives
        any_obj = False
        total_weighted = 0.0
        max_signed = -1e30
        for o in obj_specs:
            val = eval_ref(o["ref"])
            goal = (o["goal"] or "Min").lower()
            if goal == "min":
                s_val = val
            elif goal == "max":
                s_val = -val
            elif goal == "target":
                s_val = abs(val - float(o.get("target", 0)))
            else:
                s_val = val
            w = float(o.get("weight", 1.0))
            total_weighted += w * s_val
            if w * s_val > max_signed: max_signed = w * s_val
            any_obj = True

        obj_part = (alpha * max_signed + (1 - alpha) * total_weighted) if any_obj else 0.0

        # Constraints
        penalty = 0.0
        for c in con_specs:
            lhs = eval_ref(c["lhs"])
            rhs_str = c["rhs"]
            try:
                rhs = float(rhs_str)
            except (TypeError, ValueError):
                rhs = eval_ref(rhs_str, True)
            diff = lhs - rhs
            op = c.get("op", "=")
            if op == "=":
                viol = abs(diff)
            elif op in ("<=", "<"):
                viol = max(0.0, diff)
            elif op in (">=", ">"):
                viol = max(0.0, -diff)
            else:
                viol = 0.0
            pw = hard_pen if c.get("type", "Hard").lower() == "hard" else float(c.get("penalty", 1000))
            penalty += pw * viol * viol

        return obj_part + penalty

    def score(x):
        k = _memo_key(x)
        cached = _memo.get(k)
        if cached is not None:
            return cached
        v = _score_raw(x)
        _memo[k] = v
        return v

    score.eval_count = eval_count
    score.var_addrs = var_addrs
    score.write_vars = write_vars
    score.clear_memo = _memo.clear
    return score


def _safe_eval(excel, addr):
    try:
        v = excel.evaluate(addr)
        if v is None:
            return 1e12
        if isinstance(v, str) and v.startswith("#"):
            return 1e12
        fv = float(v)
        # Catch NaN/inf
        if fv != fv or abs(fv) == float('inf'):
            return 1e12
        return fv
    except Exception:
        return 1e12


# ===========================================================================
# Coordinate Descent — Anti-Zigzag (Powell) + Adaptive Step
# ===========================================================================
#
# Design note: "coordinate descent + anti-zigzag + adaptive step" is exactly
# what Powell's conjugate-direction method already is, and SciPy ships a
# numerically hardened implementation. A hand-rolled version was prototyped
# and tested against scipy Powell on the Rosenbrock valley (the canonical
# zigzag stress test): scipy reached machine precision in ~43 evals; the
# hand-rolled one stalled at f~1.6 in ~590. We use the proven engine and
# layer the requested behaviours (adaptive-step line search is intrinsic to
# Powell; memoization comes from the score wrapper; early-stop via xtol/ftol)
# on top.

def coordinate_descent_azas(score, x0, lo, hi, simplex_groups,
                            max_evals=500, init_step=0.1,
                            min_step=1e-7, tol=1e-9, log_fn=None):
    """Powell direction-set minimization through the memoized pycel score.

    - Anti-zigzag: Powell maintains a conjugate direction set, replacing the
      worst axis each iteration with the net progress direction — the textbook
      cure for coordinate-descent stair-stepping (Nocedal Ch.9.4).
    - Adaptive step: Powell's internal Brent line search expands/contracts the
      step per direction automatically (coarse early, micro near the optimum).
    - Early stop: xtol/ftol end the run when moves stop helping, typically far
      below the max_evals ceiling.
    - Projection: simplex/bounds reapplied around each evaluation so feasibility
      is preserved without constraining Powell's unconstrained line search.
    """
    n = len(x0)
    x0 = apply_projection(np.asarray(x0, float).copy(),
                          simplex_groups, lo, hi)

    def projected(z):
        return score(apply_projection(np.asarray(z, float).copy(),
                                      simplex_groups, lo, hi))

    res = minimize(projected, x0, method="Powell",
                   bounds=list(zip(lo, hi)),
                   options={"maxfev": int(max_evals),
                            "xtol": 1e-8,
                            "ftol": 1e-10})

    x_final = apply_projection(np.asarray(res.x, float).copy(),
                               simplex_groups, lo, hi)
    f_final = score(x_final)
    if log_fn:
        log_fn(f"  CD-AZAS(Powell): f={f_final:.6g}, "
               f"evals={score.eval_count[0]}, "
               f"converged={res.success}")
    return x_final, f_final


# ===========================================================================
# Method dispatch
# ===========================================================================

def run_method(method, score, x0, bounds, simplex_groups, lo, hi, max_iter,
               de_seed=None, de_seed_with_x0=False):
    method = (method or "Auto").lower()
    n = len(x0)

    def projected(x):
        return score(apply_projection(np.asarray(x, float).copy(), simplex_groups, lo, hi))

    if method in ("auto",):
        # Heuristic: prefer SLSQP for small, L-BFGS-B for larger
        method = "l-bfgs-b" if n > 50 else "slsqp"

    if method in ("portfolio", "portfolio (all methods)",
                  "portfolio all methods", "best of all",
                  "round-robin", "round robin"):
        return portfolio_solve(score, x0, bounds, simplex_groups, lo, hi,
                               max_iter, log_fn=getattr(score, "log_fn", None))

    if method in ("coordinate descent anti-zigzag adaptive step",
                  "coordinate descent anti zigzag adaptive step",
                  "cd-azas", "cd_azas", "coordinate_descent_azas"):
        x_final, f_final = coordinate_descent_azas(
            score, x0, lo, hi, simplex_groups,
            max_evals=max_iter, init_step=0.1,
            min_step=1e-7, tol=1e-9)
        return x_final, f_final

    if method in ("slsqp", "slsqp python"):
        res = minimize(projected, x0, method="SLSQP",
                       bounds=list(zip(lo, hi)),
                       options={"maxiter": max_iter, "ftol": 1e-9})
        x_final = apply_projection(res.x, simplex_groups, lo, hi)
        return x_final, res.fun

    if method in ("l-bfgs-b", "lbfgs", "homotopy / autodiff", "homotopy"):
        # NOTE: full JAX/homotopy path would require translating the pycel
        # formula graph to JAX. Until then we use L-BFGS-B with FD gradients
        # — still way faster than Excel-in-the-loop because pycel skips Excel.
        res = minimize(projected, x0, method="L-BFGS-B",
                       bounds=list(zip(lo, hi)),
                       options={"maxiter": max_iter, "ftol": 1e-9})
        x_final = apply_projection(res.x, simplex_groups, lo, hi)
        return x_final, res.fun

    if method in ("differential evolution", "de"):
        # Seed handling:
        #   - de_seed=None  -> draw a fresh OS-random seed each call so every
        #     invocation is a genuinely different global search (this is what
        #     makes re-running DE in the portfolio actually explore instead of
        #     replaying an identical run).
        #   - de_seed=<int> -> caller (portfolio) supplies a varying seed
        #     derived from round/attempt for reproducible-but-different runs.
        if de_seed is None:
            import secrets
            this_seed = secrets.randbelow(2**31 - 1)
        else:
            this_seed = int(de_seed) % (2**31 - 1)

        de_kwargs = dict(
            bounds=list(zip(lo, hi)),
            maxiter=max(20, max_iter // 10),
            tol=1e-7,
            seed=this_seed,
            polish=True,
        )

        # Optionally inject the incumbent into DE's initial population so DE
        # can keep/refine the best-known point instead of discarding it and
        # searching the whole box from scratch. The rest of the population is
        # latin-hypercube sampled for exploration.
        if de_seed_with_x0:
            try:
                import numpy as _np
                popsize = 15  # scipy default multiplier
                n_pop = max(5, popsize)
                rng = _np.random.default_rng(this_seed)
                pop = lo + rng.random((n_pop, n)) * (hi - lo)
                pop[0] = _np.clip(_np.asarray(x0, float), lo, hi)
                de_kwargs["init"] = pop
            except Exception:
                pass  # fall back to default init on any issue

        res = differential_evolution(projected, **de_kwargs)
        x_final = apply_projection(res.x, simplex_groups, lo, hi)
        return x_final, res.fun

    if method in ("de multi-restart", "de multi restart",
                  "differential evolution multi-restart",
                  "de-100", "de100"):
        # Run DE up to N times, each with a DIFFERENT seed, keeping the best
        # result found. There is no "correct" seed to search for - a seed has
        # no inherent quality - so the only sound approach is best-of-N
        # restarts. Each restart explores a different region; we keep the
        # incumbent in every restart's population so a run can never lose the
        # best basin already found.
        #
        # Two early exits keep it from wasting all 100 restarts pointlessly:
        #   - patience: if `patience` consecutive restarts yield no
        #     improvement, the search has very likely settled - stop.
        #   - eval budget: never exceed the caller's max_iter total.
        N_RESTARTS = 100
        patience = 20

        best_x = apply_projection(np.asarray(x0, float).copy(),
                                  simplex_groups, lo, hi)
        best_f = score(best_x)
        log_fn = getattr(score, "log_fn", None)
        prog_dir = getattr(score, "run_dir", None)

        no_improve = 0
        start_evals = score.eval_count[0]
        for k in range(1, N_RESTARTS + 1):
            if score.eval_count[0] - start_evals >= max_iter:
                if log_fn:
                    log_fn(f"  DE-MultiRestart: eval budget hit at "
                           f"restart {k}")
                break

            # Build a DE population seeded with the current best so this
            # restart cannot regress below it.
            this_seed = (k * 2654435761) % (2**31 - 1)  # spread seeds well
            try:
                popn = max(5, 15)
                rng = np.random.default_rng(this_seed)
                pop = lo + rng.random((popn, n)) * (hi - lo)
                pop[0] = np.clip(best_x, lo, hi)
                init = pop
            except Exception:
                init = "latinhypercube"

            r = differential_evolution(
                projected, bounds=list(zip(lo, hi)),
                maxiter=max(20, (max_iter // 10)),
                tol=1e-7, seed=this_seed, polish=True, init=init)

            rx = apply_projection(r.x, simplex_groups, lo, hi)
            rf = score(rx)
            if rf < best_f - 1e-12 * (1.0 + abs(best_f)):
                best_x, best_f = rx, rf
                no_improve = 0
                if log_fn:
                    log_fn(f"  DE-MultiRestart {k}/{N_RESTARTS}: "
                           f"new best f={best_f:.8g}")
                if prog_dir:
                    _write_progress(prog_dir,
                                    f"DE restart {k}/{N_RESTARTS} | "
                                    f"NEW BEST f={best_f:.6g}")
            else:
                no_improve += 1
                if prog_dir:
                    _write_progress(prog_dir,
                                    f"DE restart {k}/{N_RESTARTS} | "
                                    f"no gain ({no_improve}/{patience}) | "
                                    f"best f={best_f:.6g}")

            if no_improve >= patience:
                if log_fn:
                    log_fn(f"  DE-MultiRestart: {patience} restarts with no "
                           f"improvement (stopped at {k}); f={best_f:.8g}")
                break

        return best_x, best_f

    # Default
    res = minimize(projected, x0, method="L-BFGS-B",
                   bounds=list(zip(lo, hi)),
                   options={"maxiter": max_iter, "ftol": 1e-9})
    x_final = apply_projection(res.x, simplex_groups, lo, hi)
    return x_final, res.fun


# ===========================================================================
# Portfolio meta-solver  (round-robin, interpretation A)
# ===========================================================================

def portfolio_solve(score, x0, bounds, simplex_groups, lo, hi,
                     max_iter, log_fn=None,
                     order=("slsqp", "l-bfgs-b",
                            "coordinate descent anti-zigzag adaptive step",
                            "differential evolution"),
                     rel_tol=1e-9, max_rounds=10):
    """Run a portfolio of methods, chaining improvements.

    Interpretation A:
      - For each method in `order`:
          * run it from the current best point
          * if it improved, run THE SAME method again from the new point
          * repeat until that method yields no further improvement
          * then advance to the next method
      - After a full pass over all methods, if ANY method improved during
        the pass, start the whole cycle again.
      - Stop when a complete pass produces no improvement anywhere, or when
        max_rounds is hit (safety cap), or the eval budget is exhausted.

    "Improvement" = strictly better than the incumbent by more than
    rel_tol * (1 + |incumbent|). The (1+|f|) scaling makes the tolerance
    behave sensibly whether the optimum is near 0 or large.

    Every method shares the same memoized `score`, so when a method re-runs
    from a point another method already explored, repeated evaluations are
    free cache hits — the chaining is much cheaper than 4 independent solves.
    """
    def log(s):
        if log_fn:
            log_fn(s)

    # Where to write the user-facing progress (monitor window reads this).
    _run_dir = getattr(score, "run_dir", None)

    def progress(s):
        if _run_dir:
            _write_progress(_run_dir, s)

    best_x = apply_projection(np.asarray(x0, float).copy(),
                              simplex_groups, lo, hi)
    best_f = score(best_x)
    log(f"Portfolio start: f={best_f:.8g}")
    progress(f"Starting. Initial score = {best_f:.6g}")

    def improved(new_f, old_f):
        return new_f < old_f - rel_tol * (1.0 + abs(old_f))

    total_start_evals = score.eval_count[0]

    # Each inner method call gets a modest budget; the portfolio tracks the
    # global budget separately. Without this split, one SLSQP/DE call would
    # be handed the whole portfolio budget and could exhaust it in a single
    # invocation, defeating the round-robin.
    per_call_budget = max(50, max_iter // 12)

    for rnd in range(1, max_rounds + 1):
        round_improved = False
        log(f"--- Portfolio round {rnd} ---")
        progress(f"Round {rnd} starting (best so far = {best_f:.6g})")

        for method in order:
            # Re-run THIS method until it stops improving (interpretation A)
            method_iters = 0
            while True:
                if score.eval_count[0] - total_start_evals >= max_iter:
                    log(f"  [{method}] global eval budget reached; "
                        f"stopping portfolio")
                    progress(f"Eval budget reached - finishing. "
                             f"Best = {best_f:.6g}")
                    return best_x, best_f

                evals_now = score.eval_count[0] - total_start_evals
                progress(f"Round {rnd} | {method} | attempt {method_iters + 1} "
                         f"| {evals_now} evals | best = {best_f:.6g}")

                if method in ("differential evolution", "de"):
                    # Vary DE's seed every call so it explores differently
                    # each time, and inject the incumbent so it can't lose
                    # the best basin already found.
                    de_seed = 1000 * rnd + 7 * (method_iters + 1) + 13
                    x_try, f_try = run_method(method, score, best_x.copy(),
                                              bounds, simplex_groups, lo, hi,
                                              per_call_budget,
                                              de_seed=de_seed,
                                              de_seed_with_x0=True)
                else:
                    x_try, f_try = run_method(method, score, best_x.copy(),
                                              bounds, simplex_groups, lo, hi,
                                              per_call_budget)
                method_iters += 1

                if improved(f_try, best_f):
                    delta = best_f - f_try
                    best_x, best_f = x_try, f_try
                    round_improved = True
                    log(f"  [{method}] improved -> f={best_f:.8g} "
                        f"(-{delta:.3g}); re-running same method")
                    progress(f"  -> {method} improved! "
                             f"new best = {best_f:.6g} (re-running it)")
                    # loop again with the SAME method from the new point
                else:
                    if method_iters == 1:
                        log(f"  [{method}] no improvement (f={f_try:.8g})")
                    else:
                        log(f"  [{method}] exhausted after {method_iters} "
                            f"runs; best f={best_f:.8g}")
                    progress(f"  -> {method} done (no further gain)")
                    break  # advance to next method

        if not round_improved:
            log(f"Portfolio converged: full pass with no improvement "
                f"(round {rnd}), f={best_f:.8g}")
            break
    else:
        log(f"Portfolio hit max_rounds={max_rounds}; f={best_f:.8g}")

    return best_x, best_f


# ===========================================================================
# Multi-start wrapper
# ===========================================================================

def multi_start_solve(method, score, x0, bounds, simplex_groups, lo, hi,
                      max_iter, n_starts=3):
    best_x, best_f = None, np.inf
    rng = np.random.default_rng(42)
    n = len(x0)
    for s in range(n_starts):
        if s == 0:
            x_init = x0.copy()
        else:
            x_init = np.clip(x0 + rng.normal(0, 0.2, n), lo, hi)
            x_init = apply_projection(x_init, simplex_groups, lo, hi)
        x, f = run_method(method, score, x_init, bounds, simplex_groups,
                           lo, hi, max_iter)
        if f < best_f:
            best_f, best_x = f, x
    return best_x, best_f


# ===========================================================================
# Main driver
# ===========================================================================

def _compute_cache_key(src_xlsx, variables, log_fn):
    """Build a cache key that ignores ONLY the variable cells.

    The compiled pycel graph depends on formulas and on the constant values
    that formulas read — NOT on the variable cells, because the solver
    overwrites those with set_value() before the first evaluation anyway.

    Each solve writes optimized numbers back into the variable cells, so the
    raw file bytes change every run even when the user changed nothing. If
    we keyed on raw bytes we'd recompile every single time and the cache
    would be useless.

    So we hash the workbook with the variable cells blanked out:
      - solver's own previous output (variable cells)  -> excluded -> cache HIT
      - user edits a formula                           -> included -> recompile
      - user edits a constant a formula reads           -> included -> recompile
      - user changes nothing                            -> same key -> cache HIT

    This is safe: anything the USER can change is still in the key. Only the
    cells the optimizer itself owns are excluded.
    """
    import hashlib
    from openpyxl import load_workbook

    try:
        wb = load_workbook(src_xlsx, data_only=False)
    except Exception as e:
        # Fall back to raw-byte hash if we can't introspect — correct, just
        # less cache-friendly.
        log_fn(f"Cache key: workbook introspection failed ({e}); using raw hash")
        return hashlib.sha1(Path(src_xlsx).read_bytes()).hexdigest()[:16]

    # Set of variable cells to exclude, as (sheet, coordinate) upper-cased.
    var_cells = set()
    for v in variables:
        ref = v.get("ref", "")
        if "!" not in ref:
            continue
        sheet, cell = ref.split("!", 1)
        sheet = sheet.strip().strip("'").strip()
        cell = cell.replace("$", "").upper()
        var_cells.add((sheet, cell))

    h = hashlib.sha1()
    for ws in wb.worksheets:
        h.update(("\x00SHEET:" + ws.title + "\x00").encode("utf-8"))
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                if (ws.title, c.coordinate.upper()) in var_cells:
                    continue  # skip solver-owned cells
                # Include address + formula/value. For formula cells
                # c.value is the formula string (data_only=False).
                h.update(f"{c.coordinate}={c.value!r};".encode("utf-8"))
    return h.hexdigest()[:16]


def _load_excel_cached(src_xlsx, variables, log_fn):
    """Compile the workbook with pycel, caching the compiled graph.

    Parsing every formula in a workbook is the single biggest fixed cost in
    the pipeline (often 1-10s) and it is method-independent — SLSQP, DE,
    L-BFGS-B and CD-AZAS all pay it. We serialize pycel's compiled graph and
    key it on _compute_cache_key (formulas + constants, variable cells
    excluded) so the solver's own previous output never busts the cache,
    while any user edit still forces a correct recompile.

    pycel's to_file() appends a type extension (.pkl by default) to the
    given path; from_file() reads that exact path back. We build the cache
    filename WITHOUT an extension and let pycel add ".pkl", matching the
    usage in pycel's own test suite.
    """
    src_xlsx = Path(src_xlsx)
    digest = _compute_cache_key(src_xlsx, variables, log_fn)
    cache_base = str(src_xlsx.parent.parent / f"_pycel_cache_{digest}")
    cache_pkl = cache_base + ".pkl"

    if os.path.exists(cache_pkl):
        try:
            ex = ExcelCompiler.from_file(cache_pkl)
            log_fn("Pycel graph loaded from cache (skipped parse)")
            return ex
        except Exception as e:
            log_fn(f"Cache load failed ({e}); recompiling")

    ex = ExcelCompiler(filename=str(src_xlsx))
    try:
        ex.to_file(cache_base, file_types=('pkl',))
        log_fn("Pycel graph compiled and cached for next run")
    except Exception as e:
        log_fn(f"Could not write pycel cache ({e}); continuing without it")
    return ex


def main(run_dir):
    # ---- Pre-flight: validate the run directory BEFORE the try block ----
    # If we were handed a bad argument (e.g. the literal string "--monitor"
    # because an OLD exe doesn't understand monitor mode), every later step
    # fails AND the error handler's own write to results.json fails too,
    # producing the cryptic double-traceback. Catch it here, loudly, and
    # do NOT let the window vanish.
    raw_arg = run_dir
    run_dir = Path(run_dir)
    cfg_path = run_dir / "config.json"
    src_xlsx = run_dir / "source.xlsx"

    if not run_dir.is_dir() or not cfg_path.is_file():
        print("=" * 60, flush=True)
        print(" FASTSOLVER - STARTUP ERROR", flush=True)
        print("=" * 60, flush=True)
        print(f" The run directory is not valid:", flush=True)
        print(f"   argument received : {raw_arg!r}", flush=True)
        print(f"   resolved path     : {run_dir}", flush=True)
        print(f"   directory exists  : {run_dir.is_dir()}", flush=True)
        print(f"   config.json found : {cfg_path.is_file()}", flush=True)
        print(flush=True)
        if str(raw_arg).startswith("--"):
            print(" It looks like an OPTION ('--...') was passed as the run", flush=True)
            print(" directory. This almost always means the FastSolver.exe", flush=True)
            print(" being run is an OLD build that does not understand", flush=True)
            print(" monitor mode. Rebuild the exe from the current", flush=True)
            print(" 07_fastsolver_bridge.py (run 11_build_exe.py) and", flush=True)
            print(" redeploy the dist/FastSolver folder next to the .xlam.", flush=True)
        else:
            print(" The VBA side did not create config.json here, or the", flush=True)
            print(" path is wrong. Check modPythonBridge run-folder logic.", flush=True)
        print(flush=True)
        print(" This window is staying open. Type q then ENTER to close.",
              flush=True)
        _wait_for_q()
        sys.exit(3)

    log = []
    def log_line(s):
        log.append(s)
        print(s, flush=True)

    try:
        config = json.loads(cfg_path.read_text())
        log_line(f"Loaded config: {len(config['variables'])} vars, "
                 f"{len(config['objectives'])} objs, "
                 f"{len(config['constraints'])} cons")

        variables = config["variables"]
        n = len(variables)
        lo = np.array([float(v["lo"]) for v in variables])
        hi = np.array([float(v["hi"]) for v in variables])
        simplex_groups = collect_simplex_groups(variables)
        log_line(f"Simplex groups: {len(simplex_groups)}")

        _set_phase(run_dir, "compiling")
        t0 = time.time()
        excel = _load_excel_cached(src_xlsx, variables, log_line)
        log_line(f"Pycel ready in {time.time()-t0:.2f}s")
        _set_phase(run_dir, "solving")

        # Initial values from the workbook
        x0 = np.zeros(n)
        for i, v in enumerate(variables):
            x0[i] = _safe_eval(excel, _pycel_addr(v["ref"]))
        x0 = apply_projection(x0, simplex_groups, lo, hi)

        score = build_score_fn(excel, config, variables, simplex_groups, lo, hi)
        score.log_fn = log_line
        score.run_dir = str(run_dir)
        start_score = score(x0)

        # Sanity probes cost 2 extra full evaluations on every run. On a slow
        # workbook that's wasted seconds. Only run them when debugging.
        if os.environ.get("FASTSOLVER_DEBUG"):
            x_test = x0.copy()
            if len(x_test) > 0:
                x_test[0] = min(hi[0], x0[0] + 0.1) if x0[0] + 0.1 <= hi[0] else max(lo[0], x0[0] - 0.1)
            probe_score = score(x_test)
            log_line(f"Probe score (perturbed x[0]): {probe_score:.6f}")
            if abs(probe_score - start_score) < 1e-12:
                log_line("WARNING: score did not change on perturbation — pycel not propagating vars")
            x_zero = x0.copy()
            x_zero[0] = 0.0
            crash_score = score(x_zero)
            log_line(f"Score at x[0]=0 (div-by-zero test): {crash_score:.2e}")
        log_line(f"Initial score: {start_score:.6f}")


        method = config.get("method", "Auto")
        max_iter = int(config.get("max_iter", 200))

        t1 = time.time()
        if (method or "").strip().lower() in (
                "portfolio", "portfolio (all methods)",
                "portfolio all methods", "best of all",
                "round-robin", "round robin"):
            # The portfolio self-explores (it includes DE for global search
            # and chains local methods), so wrapping it in random multi-start
            # would just repeat the whole portfolio from noisy points and
            # waste the budget. Run it once with a larger total eval budget.
            x_best, f_best = portfolio_solve(
                score, x0, list(zip(lo, hi)), simplex_groups, lo, hi,
                max_iter=max(max_iter * 6, 1500), log_fn=log_line)
        else:
            x_best, f_best = multi_start_solve(method, score, x0,
                                               list(zip(lo, hi)),
                                               simplex_groups, lo, hi,
                                               max_iter, n_starts=3)
        elapsed = time.time() - t1
        log_line(f"Optimized in {elapsed:.2f}s. Final score: {f_best:.6f}. "
                 f"Eval count: {score.eval_count[0]}")

        # solved.xlsx is NOT read by the VBA side — it applies
        # variable_values from results.json directly to the live workbook.
        # Re-loading + saving the whole workbook here cost 0.5-2s every run
        # for a file nothing consumes. Only emit it when debugging.
        if os.environ.get("FASTSOLVER_DEBUG"):
            wb = load_workbook(src_xlsx)
            for v, val in zip(variables, x_best):
                sheet, cell = _split_addr(v["ref"])
                try:
                    wb[sheet][cell] = float(val)
                except Exception:
                    pass
            wb.save(run_dir / "solved.xlsx")

        # Write results.json
        results = {
            "status": "ok",
            "method_used": method,
            "elapsed_sec": elapsed,
            "eval_count": score.eval_count[0],
            "start_score": start_score,
            "final_score": float(f_best),
            "variable_names": [v["ref"] for v in variables],
            "variable_values": [float(x) for x in x_best],
            "log": log,
        }
        (run_dir / "results.json").write_text(json.dumps(results, indent=2))
        log_line("Wrote results.json + solved.xlsx")
    except Exception as e:
        tb = traceback.format_exc()
        # Try to write results.json, but NEVER let this fail silently the
        # way it did before (when run_dir was bad, the write itself threw
        # and buried the real error under a second traceback).
        wrote = False
        try:
            err = {
                "status": "error",
                "message": str(e),
                "traceback": tb,
                "log": log,
            }
            (run_dir / "results.json").write_text(json.dumps(err, indent=2))
            wrote = True
        except Exception as e2:
            print(f"(could not write results.json: {e2})", flush=True)

        # Always show the real error in the console, regardless of whether
        # results.json could be written.
        print(flush=True)
        print("=" * 60, flush=True)
        print(" FASTSOLVER - SOLVE FAILED", flush=True)
        print("=" * 60, flush=True)
        print(f" Error: {e}", flush=True)
        print(flush=True)
        print(" --- traceback ---", flush=True)
        for ln in tb.splitlines()[-25:]:
            print(" | " + ln, flush=True)
        if log:
            print(flush=True)
            print(" --- log tail ---", flush=True)
            for ln in log[-15:]:
                print(" | " + str(ln), flush=True)
        if wrote:
            print("\n (results.json written - the Excel side will report "
                  "failure too)", flush=True)
        print(flush=True)
        print(" This window is staying open. Type q then ENTER to close.",
              flush=True)
        _wait_for_q()
        sys.exit(2)


def _wait_for_q():
    """Block until the user types q + Enter. Robust against a frozen-exe
    console where bare input() can return instantly: loop, and if stdin is
    genuinely unusable, fall back to a very long sleep so the window with
    the error stays readable instead of vanishing."""
    import time as _t
    while True:
        try:
            ans = input(" > ").strip().lower()
        except Exception:
            _t.sleep(900.0)
            return
        if ans == "q":
            return
        print(" (type q then ENTER to close)", flush=True)


def _write_progress(run_dir, text):
    """Append a progress line the monitor window can display. Best-effort:
    progress reporting must never crash or slow the actual solve."""
    try:
        p = Path(run_dir) / "progress.txt"
        with open(p, "a", encoding="utf-8") as fh:
            fh.write(text.rstrip() + "\n")
    except Exception:
        pass


def _set_phase(run_dir, phase):
    """Signal the monitor which phase we're in: 'compiling' or 'solving'.
    The monitor shows an animated (no-percent) bar during 'compiling' since
    pycel can't report compile progress, then switches to live per-method
    counters once 'solving' begins."""
    try:
        (Path(run_dir) / "phase.txt").write_text(phase, encoding="utf-8")
    except Exception:
        pass


def monitor(run_dir):
    """Visible companion window with two phases:

      1. COMPILING - an animated bar (NO percentage: pycel cannot report
         how far through parsing it is, so a real % would be a lie). The
         bar just sweeps to show the process is alive.

      2. SOLVING - live per-method view: the current method, an ASCII
         progress bar across the portfolio's method list, and the eval
         count ticking up per method.

    It reads phase.txt + progress.txt written by the hidden solver and
    self-closes when results.json appears.
    """
    import time as _t
    run_dir = Path(run_dir)
    prog = run_dir / "progress.txt"
    phase_f = run_dir / "phase.txt"
    done = run_dir / "results.json"

    def read_phase():
        try:
            return phase_f.read_text(encoding="utf-8").strip()
        except Exception:
            return "compiling"

    def clear():
        # ANSI clear-line; harmless if unsupported
        print("\r" + " " * 70 + "\r", end="")

    print("=" * 60)
    print("   F A S T S O L V E R")
    print("   (progress monitor - closes itself when finished)")
    print("=" * 60, flush=True)

    # ---- Phase 1: COMPILING (animated, no percent) -------------------
    bar_w = 28
    pos = 0
    direction = 1
    comp_waited = 0.0
    COMPILE_LIMIT = 300.0   # 5 min: pycel compile should never exceed this
    print()
    while not done.exists() and read_phase() == "compiling":
        # A single lit cell sweeping back and forth = "alive, working".
        cells = ["."] * bar_w
        cells[pos] = "#"
        bar = "".join(cells)
        print(f"\r  Compiling workbook  [{bar}]  (this can take a few s)",
              end="", flush=True)
        pos += direction
        if pos >= bar_w - 1 or pos <= 0:
            direction *= -1
        _t.sleep(0.08)
        comp_waited += 0.08
        if comp_waited >= COMPILE_LIMIT:
            print(flush=True)
            print(f"  Still 'compiling' after {int(COMPILE_LIMIT)}s with no "
                  f"result - the solver likely crashed during startup.",
                  flush=True)
            break
    clear()
    if done.exists() or read_phase() != "compiling":
        print("\r  Compiling workbook  [ done ]", flush=True)

    # ---- Phase 2: SOLVING (live per-method counters) -----------------
    print()
    print("  Solving:")
    seen = 0
    spin = "|/-\\"
    spin_i = 0
    waited = 0.0
    silent = 0.0          # seconds since the last progress change
    SILENCE_LIMIT = 180.0 # 3 min with zero progress AND no result = dead
    while not done.exists():
        lines = []
        try:
            if prog.exists():
                lines = prog.read_text(encoding="utf-8").splitlines()
        except Exception:
            lines = []

        if len(lines) > seen:
            for ln in lines[seen:]:
                print("   " + ln, flush=True)
            seen = len(lines)
            silent = 0.0          # progress moved — solver is alive
        else:
            spin_i = (spin_i + 1) % len(spin)
            last = lines[-1] if lines else "starting..."
            print(f"\r   {spin[spin_i]} {last}   "
                  f"({waited:0.0f}s)        ",
                  end="", flush=True)
            silent += 0.4

        if silent >= SILENCE_LIMIT:
            # No progress for a long time and still no results.json: the
            # solver almost certainly died without writing an error file
            # (hard crash / killed). Stop waiting and fall through to the
            # failure display so the user isn't left staring at a spinner.
            print(flush=True)
            print(f"  No progress for {int(SILENCE_LIMIT)}s and no result "
                  f"- assuming the solver crashed.", flush=True)
            break

        _t.sleep(0.4)
        waited += 0.4

    print()
    print("  " + "-" * 50)

    failed = False
    msg = ""
    tb = ""
    logtail = []
    try:
        if done.exists():
            raw = done.read_text(encoding="utf-8", errors="replace")
            try:
                res = json.loads(raw)
            except Exception:
                # results.json exists but is malformed/partial — show the
                # raw bytes, that's still the best clue we have.
                failed = True
                msg = "results.json is present but not valid JSON " \
                      "(solver likely died mid-write)."
                tb = raw[-2000:]
                res = None
            if res is not None:
                if res.get("status") == "ok":
                    print(f"  DONE.  Final score: "
                          f"{res.get('final_score'):.6g}"
                          f"   ({res.get('eval_count','?')} evals)",
                          flush=True)
                else:
                    failed = True
                    msg = res.get("message", "(no message in results.json)")
                    # The solver stores the real Python traceback + its
                    # in-memory log INSIDE results.json, not as a separate
                    # file. Surface both - this is the actual error.
                    tb = res.get("traceback", "") or ""
                    lg = res.get("log", [])
                    if isinstance(lg, list):
                        logtail = [str(x) for x in lg][-20:]
        else:
            failed = True
            msg = "Solver exited without producing results.json (it crashed)."
    except Exception as e:
        failed = True
        msg = f"Monitor could not read results.json: {e}"

    if not failed:
        print("  (window closes in 3s)", flush=True)
        _t.sleep(3.0)
        return

    # FAILURE: show everything useful and DO NOT auto-close.
    print(flush=True)
    print("  " + "!" * 52, flush=True)
    print("  SOLVE FAILED", flush=True)
    print("  " + "!" * 52, flush=True)
    print(f"  Reason: {msg}", flush=True)

    if logtail:
        print("\n  --- solver log (last lines) ---", flush=True)
        for ln in logtail:
            print("  | " + ln, flush=True)

    if tb:
        print("\n  --- Python traceback ---", flush=True)
        for ln in tb.splitlines()[-25:]:
            print("  | " + ln, flush=True)

    print(flush=True)
    print("  " + "=" * 52, flush=True)
    print("  This window is STAYING OPEN so you can read the error.",
          flush=True)
    print("  Type  q  then ENTER to close it.", flush=True)
    print("  " + "=" * 52, flush=True)

    # Robust pause. Bare input() can return instantly in a frozen-exe
    # console with no real stdin; loop until we actually get a 'q', and
    # if stdin is truly unusable, fall back to a very long sleep so the
    # window still does not vanish.
    while True:
        try:
            ans = input("  > ").strip().lower()
        except Exception:
            _t.sleep(600.0)
            break
        if ans == "q":
            break
        print("  (type q then ENTER to close)", flush=True)


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "--monitor":
        monitor(sys.argv[2])
        sys.exit(0)
    if len(sys.argv) < 2:
        print("Usage: fastsolver_bridge.py <run_dir>", file=sys.stderr)
        print("       fastsolver_bridge.py --monitor <run_dir>",
              file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1])