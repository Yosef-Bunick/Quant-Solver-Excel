"""
fastsolver_bridge.py — Python-side entry point for the .xlam add-in.

Called as:
    FastSolver.exe <run_dir>

Where <run_dir> contains:
    config.json    — problem spec exported by VBA
    source.xlsx    — copy of the user's workbook

Produces:
    results.json   — final score + variable name/value arrays
    solved.xlsx    — workbook copy with optimal values written in

Implements method dispatch:
    "Auto"                  → routes by problem size/structure
    "SLSQP Python"          → SciPy SLSQP via pycel evaluation
    "L-BFGS-B"              → scaled to large var counts
    "Differential Evolution"→ global, derivative-free, pycel-eval
    "Homotopy / AutoDiff"   → JAX + smoothing — placeholder for full
                              graph→JAX rewrite; falls back to L-BFGS-B
                              over pycel-eval with finite-diff today
"""

from __future__ import annotations
import json
import os
import sys
import time
import traceback
from pathlib import Path
import numpy as np

# ---------------------------------------------------------------------------
# Defined-name handling for pycel 1.0b30
# ---------------------------------------------------------------------------
# pycel 1.0b30 mis-parses workbook/sheet defined names embedded in formulas as
# structured-table names ("Table Name not found: LogLogScale") and cannot
# disambiguate sheet-scoped vs workbook-scoped duplicates. The fix is to
# expand supported defined names to their target cell/range refs BEFORE pycel
# compiles, then strip any remaining #REF! names from the Name Manager table.


def _dn_container_values(container):
    """Return DefinedName-like objects from old/new openpyxl containers."""
    if container is None:
        return []

    try:
        vals = list(container.definedName)
    except Exception:
        try:
            vals = list(container.values())
        except Exception:
            try:
                vals = list(container)
            except Exception:
                vals = []

    out = []
    for v in vals:
        if isinstance(v, (list, tuple, set)):
            out.extend(v)
        else:
            out.append(v)
    return out


def _dn_text(d):
    """openpyxl has used attr_text/value/text across versions."""
    for attr in ("attr_text", "value", "text"):
        try:
            v = getattr(d, attr, None)
        except Exception:
            v = None
        if v:
            return str(v)
    return None


def _dn_name(d):
    try:
        return getattr(d, "name", None) or getattr(d, "localName", None)
    except Exception:
        return None


def _sheet_for_local_id(wb, local_sheet_id):
    if local_sheet_id is None:
        return None
    try:
        return wb.worksheets[int(local_sheet_id)].title
    except Exception:
        return None


def _quote_sheet(sheet):
    if sheet is None:
        return None
    escaped = str(sheet).replace("'", "''")
    return "'" + escaped + "'"


def _iter_defined_names_from_xlsx_xml(xlsx_path):
    """Yield defined names by reading xl/workbook.xml directly.

    Version-independent fallback. Some openpyxl releases expose sheet-scoped
    duplicate names inconsistently; Excel's Name Manager entries are always in
    workbook.xml as <definedName>, so we read the XML too and union results.
    """
    if not xlsx_path:
        return
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            raw = zf.read("xl/workbook.xml")
        root = ET.fromstring(raw)
    except Exception:
        return

    sheet_names = []
    try:
        for sh in root.findall(".//{*}sheets/{*}sheet"):
            sheet_names.append(sh.attrib.get("name", ""))
    except Exception:
        sheet_names = []

    try:
        nodes = root.findall(".//{*}definedNames/{*}definedName")
    except Exception:
        nodes = []

    for node in nodes:
        name = node.attrib.get("name")
        if not name:
            continue
        # Skip built-in names such as _xlnm.Print_Area.
        if str(name).startswith("_xlnm."):
            continue
        target = "".join(node.itertext()).strip()
        if not target:
            continue
        sheet = None
        local_id = node.attrib.get("localSheetId")
        if local_id is not None:
            try:
                sheet = sheet_names[int(local_id)]
            except Exception:
                sheet = None
        yield str(name), str(target), sheet


def _iter_defined_names(wb, xlsx_path=None):
    """Yield (name, target, scope_sheet_or_None) from openpyxl + raw XML."""
    seen = set()

    def emit(name, target, sheet):
        if not name or not target:
            return None
        key = (str(name).casefold(),
               _strip_sheet_quotes(sheet).casefold() if sheet else None,
               str(target).strip())
        if key in seen:
            return None
        seen.add(key)
        return str(name), str(target), sheet

    for d in _dn_container_values(getattr(wb, "defined_names", None)):
        name = _dn_name(d)
        target = _dn_text(d)
        if not name or not target:
            continue
        sheet = _sheet_for_local_id(wb, getattr(d, "localSheetId", None))
        row = emit(name, target, sheet)
        if row:
            yield row

    for ws in getattr(wb, "worksheets", []):
        for d in _dn_container_values(getattr(ws, "defined_names", None)):
            name = _dn_name(d)
            target = _dn_text(d)
            if not name or not target:
                continue
            row = emit(name, target, ws.title)
            if row:
                yield row

    for name, target, sheet in _iter_defined_names_from_xlsx_xml(xlsx_path):
        row = emit(name, target, sheet)
        if row:
            yield row


# ---------------------------------------------------------------------------
# Formula rewriter for defined names (pycel 1.0b30)
# ---------------------------------------------------------------------------


def _strip_sheet_quotes(sheet):
    sheet = str(sheet or "").strip()
    if len(sheet) >= 2 and sheet[0] == "'" and sheet[-1] == "'":
        sheet = sheet[1:-1].replace("''", "'")
    return sheet.strip()


def _formula_ref_from_defined_name_target(target):
    """Return a formula-safe replacement for a defined-name target.

    Names point at one cell/range, e.g. "'LogLog Fitting'!$B$33". Broken
    names (#REF!) are deliberately skipped.
    """
    if target is None:
        return None
    t = str(target).strip()
    if not t:
        return None

    if t.startswith("="):
        t = t[1:].strip()

    if "#REF!" in t.upper():
        return None

    if "[" in t and "]" in t:
        return None

    if "!" in t and "," not in t:
        sheet, ref = t.rsplit("!", 1)
        sheet = _strip_sheet_quotes(sheet)
        if not sheet or not ref:
            return None
        return f"{_quote_sheet(sheet)}!{ref.strip()}"

    if any(op in t for op in ("+", "-", "*", "/", "^", "&", "=", "<", ">")):
        return f"({t})"
    return t


def _build_name_rewrite_maps(wb, xlsx_path=None):
    """Build workbook/local defined-name maps for formula rewriting."""
    workbook = {}
    local = {}
    skipped = []

    for name, target, sheet in _iter_defined_names(wb, xlsx_path):
        repl = _formula_ref_from_defined_name_target(target)
        if not repl:
            skipped.append((sheet or "WORKBOOK", str(name), str(target)))
            continue
        key = str(name).casefold()
        if sheet:
            skey = _strip_sheet_quotes(sheet).casefold()
            local.setdefault(skey, {})[key] = repl
        else:
            workbook[key] = repl

    return {"workbook": workbook, "local": local, "skipped": skipped}


def _lookup_bare_defined_name(name_maps, current_sheet, name):
    if not name_maps or not name:
        return None
    nkey = str(name).casefold()
    skey = _strip_sheet_quotes(current_sheet).casefold() if current_sheet else None
    if skey:
        hit = name_maps.get("local", {}).get(skey, {}).get(nkey)
        if hit is not None:
            return hit
    return name_maps.get("workbook", {}).get(nkey)


def _lookup_qualified_defined_name(name_maps, sheet, name):
    """Resolve a sheet-qualified defined-name reference (e.g.
    'Lomax Fitting'!LomaxTruncation).

    A sheet-qualified reference may point at EITHER a sheet-scoped local
    name OR a workbook-scoped name (Excel allows both spellings). pycel
    1.0b30 crashes on any defined name it has not pre-expanded, so we must
    resolve both: try the sheet-local map first, then fall back to the
    workbook map. Without the fallback, workbook-scoped names written in
    qualified form (the common case for cross-sheet objective chains) slip
    through unrewritten and pycel raises
    "'DefinedNameDict' object has no attribute 'definedName'".
    """
    if not name_maps or not sheet or not name:
        return None
    skey = _strip_sheet_quotes(sheet).casefold()
    nkey = str(name).casefold()
    hit = name_maps.get("local", {}).get(skey, {}).get(nkey)
    if hit is not None:
        return hit
    return name_maps.get("workbook", {}).get(nkey)


def _is_excel_name_start(ch):
    return bool(ch) and (ch == "_" or ch == "\\" or ch.isalpha())


def _is_excel_name_char(ch):
    return bool(ch) and (ch == "_" or ch == "\\" or ch == "." or ch.isalnum())


def _is_unquoted_sheet_char(ch):
    return bool(ch) and (ch == "_" or ch == "." or ch.isalnum())


def _rewrite_formula_segment_defined_names(segment, current_sheet, name_maps, hits=None):
    """Rewrite defined names in a formula segment that is outside strings."""
    if not segment or not name_maps:
        return segment

    n = len(segment)
    out = []
    i = 0

    while i < n:
        ch = segment[i]

        # Quoted sheet-qualified local name:  'Sheet 1'!localName
        if ch == "'":
            j = i + 1
            sheet_chars = []
            while j < n:
                if segment[j] == "'":
                    if j + 1 < n and segment[j + 1] == "'":
                        sheet_chars.append("'")
                        j += 2
                        continue
                    break
                sheet_chars.append(segment[j])
                j += 1
            if j < n and segment[j] == "'" and j + 1 < n and segment[j + 1] == "!":
                k = j + 2
                if k < n and _is_excel_name_start(segment[k]):
                    l = k + 1
                    while l < n and _is_excel_name_char(segment[l]):
                        l += 1
                    sheet = "".join(sheet_chars)
                    name = segment[k:l]
                    repl = _lookup_qualified_defined_name(name_maps, sheet, name)
                    if repl is not None:
                        if hits is not None:
                            hits[name] = hits.get(name, 0) + 1
                        out.append(repl)
                        i = l
                        continue
            out.append(ch)
            i += 1
            continue

        # Unquoted sheet-qualified local name:  Sheet1!localName
        if _is_unquoted_sheet_char(ch):
            j = i + 1
            while j < n and _is_unquoted_sheet_char(segment[j]):
                j += 1
            if j < n and segment[j] == "!":
                k = j + 1
                if k < n and _is_excel_name_start(segment[k]):
                    l = k + 1
                    while l < n and _is_excel_name_char(segment[l]):
                        l += 1
                    sheet = segment[i:j]
                    name = segment[k:l]
                    repl = _lookup_qualified_defined_name(name_maps, sheet, name)
                    if repl is not None:
                        if hits is not None:
                            hits[name] = hits.get(name, 0) + 1
                        out.append(repl)
                        i = l
                        continue
                out.append(segment[i:j])
                i = j
                continue

        # Bare workbook/local name:  LogLogShape, truncation, etc.
        if _is_excel_name_start(ch):
            j = i + 1
            while j < n and _is_excel_name_char(segment[j]):
                j += 1
            token = segment[i:j]
            repl = _lookup_bare_defined_name(name_maps, current_sheet, token)
            if repl is not None:
                if hits is not None:
                    hits[token] = hits.get(token, 0) + 1
                out.append(repl)
            else:
                out.append(token)
            i = j
            continue

        out.append(ch)
        i += 1

    return "".join(out)


def _rewrite_formula_defined_names(formula, current_sheet, name_maps, hits=None):
    """Rewrite defined names in formula text, skipping double-quoted strings."""
    if not isinstance(formula, str) or not formula.startswith("=") or not name_maps:
        return formula

    out = []
    buf = []
    i = 0
    n = len(formula)
    in_string = False

    while i < n:
        ch = formula[i]
        if ch == '"':
            if in_string:
                if i + 1 < n and formula[i + 1] == '"':
                    out.append(ch)
                    out.append(formula[i + 1])
                    i += 2
                    continue
                in_string = False
                out.append(ch)
                i += 1
                continue
            if buf:
                seg = "".join(buf)
                out.append(_rewrite_formula_segment_defined_names(seg, current_sheet, name_maps, hits))
                buf = []
            in_string = True
            out.append(ch)
            i += 1
            continue

        if in_string:
            out.append(ch)
        else:
            buf.append(ch)
        i += 1

    if buf:
        seg = "".join(buf)
        out.append(_rewrite_formula_segment_defined_names(seg, current_sheet, name_maps, hits))

    return "".join(out)


def _resolve_defined_name_reference(ref, current_sheet, name_maps):
    """Resolve a config ref that is itself a defined name, if possible."""
    if not ref or not name_maps:
        return None
    s = str(ref).strip()
    if s.startswith("="):
        return None

    if s.startswith("'") and "!" in s:
        sheet_part, name_part = s.rsplit("!", 1)
        name_part = name_part.strip().replace("$", "")
        if name_part and _is_excel_name_start(name_part[0]) and all(_is_excel_name_char(c) for c in name_part):
            return _lookup_qualified_defined_name(name_maps, sheet_part, name_part)
        return None

    if "!" in s:
        sheet_part, name_part = s.rsplit("!", 1)
        name_part = name_part.strip().replace("$", "")
        if name_part and _is_excel_name_start(name_part[0]) and all(_is_excel_name_char(c) for c in name_part):
            return _lookup_qualified_defined_name(name_maps, sheet_part, name_part)
        return None

    token = s.replace("$", "")
    if token and _is_excel_name_start(token[0]) and all(_is_excel_name_char(c) for c in token):
        return _lookup_bare_defined_name(name_maps, current_sheet, token)
    return None


def _build_name_maps_only(src_xlsx, log_fn=None):
    """Build just the defined-name maps, skipping the formula rewrite + save.

    Used on a cache HIT: the compiled pycel graph is reused from disk, but
    runtime ref resolution still needs name_maps. This does only the cheap
    half of _rewrite_workbook_formulas_for_pycel -- one workbook open and the
    map build -- and never iterates/rewrites cells or re-saves the workbook.
    """
    from openpyxl import load_workbook

    src_xlsx = Path(src_xlsx)
    try:
        wb = load_workbook(src_xlsx, data_only=False, keep_links=False)
    except Exception as e:
        if log_fn:
            log_fn(f"Name-map build skipped: could not open workbook ({e})")
        return {"workbook": {}, "local": {}, "skipped": []}

    name_maps = _build_name_rewrite_maps(wb, src_xlsx)
    if log_fn:
        n_wb = len(name_maps.get("workbook", {}))
        n_local = sum(len(v) for v in name_maps.get("local", {}).values())
        log_fn(f"Name maps rebuilt for cached graph: "
               f"{n_wb} workbook names, {n_local} sheet-local names")
    return name_maps


def _rewrite_workbook_formulas_for_pycel(src_xlsx, log_fn=None):
    """Create a pycel-safe workbook copy with defined names expanded.

    Returns (path_to_compile, name_maps). The original source.xlsx is never
    modified; only a temporary workbook in the run directory is saved.
    """
    from openpyxl import load_workbook

    src_xlsx = Path(src_xlsx)
    try:
        wb = load_workbook(src_xlsx, data_only=False, keep_links=False)
    except Exception as e:
        if log_fn:
            log_fn(f"Defined-name formula rewrite skipped: could not open workbook ({e})")
        return src_xlsx, {"workbook": {}, "local": {}, "skipped": []}

    name_maps = _build_name_rewrite_maps(wb, src_xlsx)
    if log_fn:
        n_wb = len(name_maps.get("workbook", {}))
        n_local = sum(len(v) for v in name_maps.get("local", {}).values())
        log_fn(f"Defined-name formula rewrite: discovered {n_wb} workbook names and {n_local} sheet-local names")
    hits = {}
    changed = 0
    samples = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    nv = _rewrite_formula_defined_names(v, ws.title, name_maps, hits)
                    if nv != v:
                        cell.value = nv
                        changed += 1
                        if len(samples) < 3:
                            samples.append(f"{ws.title}!{cell.coordinate}: {v} -> {nv}")

    if log_fn:
        if changed:
            top_hits = ", ".join(f"{k}={v}" for k, v in sorted(hits.items(), key=lambda kv: (-kv[1], kv[0]))[:8])
            log_fn(f"Defined-name formula rewrite: {changed} formula cells rewritten" + (f" ({top_hits})" if top_hits else ""))
            if os.environ.get("FASTSOLVER_DEBUG"):
                for s in samples:
                    log_fn("  rewrite sample: " + s[:500])
        else:
            log_fn("Defined-name formula rewrite: no formula cells needed rewriting")
        skipped = name_maps.get("skipped", [])
        if skipped and os.environ.get("FASTSOLVER_DEBUG"):
            shown = ", ".join(f"{scope}!{name}->{target}" for scope, name, target in skipped[:8])
            log_fn(f"Defined-name formula rewrite: skipped unsupported/broken names: {shown}")

    if not changed:
        return src_xlsx, name_maps

    out = src_xlsx.with_name("_fastsolver_pycel_formula_rewrite.xlsx")
    try:
        wb.save(out)
        return out, name_maps
    except Exception as e:
        if log_fn:
            log_fn(f"Defined-name formula rewrite save failed ({e}); using original workbook")
        return src_xlsx, name_maps


# ---------------------------------------------------------------------------
# Defined-name sanitizer: strip #REF! names from the Name Manager table
# ---------------------------------------------------------------------------


def _iter_defined_name_entries(wb):
    """Yield (container, defined_name_obj, name, target, scope_sheet_or_None)."""
    container = getattr(wb, "defined_names", None)
    for d in _dn_container_values(container):
        name = _dn_name(d)
        target = _dn_text(d)
        if not name or not target:
            continue
        sheet = _sheet_for_local_id(wb, getattr(d, "localSheetId", None))
        yield container, d, str(name), str(target), sheet

    for ws in getattr(wb, "worksheets", []):
        container = getattr(ws, "defined_names", None)
        for d in _dn_container_values(container):
            name = _dn_name(d)
            target = _dn_text(d)
            if not name or not target:
                continue
            yield container, d, str(name), str(target), ws.title


def _remove_defined_name_from_container(container, d, name):
    """Best-effort deletion across old/new openpyxl defined-name containers."""
    if container is None:
        return False

    try:
        if hasattr(container, "pop"):
            container.pop(name, None)
            return True
    except Exception:
        pass

    try:
        del container[name]
        return True
    except Exception:
        pass

    try:
        if hasattr(container, "definedName"):
            container.definedName.remove(d)
            return True
    except Exception:
        pass

    return False


def _defined_name_single_cell_value(wb, target):
    """Return a single-cell defined-name target's value.

    Returns None when the target is unsupported, missing, or blank (so the
    duplicate-name pass can drop a definition pointing at an empty cell when a
    sibling points at a usable one). A range target counts as non-blank.
    """
    try:
        from openpyxl.utils.cell import range_boundaries
        t = str(target or "").strip()
        if t.startswith("="):
            t = t[1:].strip()
        if "#REF!" in t.upper() or "," in t or "!" not in t:
            return None
        sheet, ref = t.rsplit("!", 1)
        sheet = _strip_sheet_quotes(sheet)
        ref = ref.replace("$", "")
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        if min_col != max_col or min_row != max_row:
            return "__RANGE__"
        if sheet not in wb.sheetnames:
            return None
        return wb[sheet].cell(min_row, min_col).value
    except Exception:
        return None


def _sanitize_defined_names_for_pycel(src_xlsx, log_fn=None):
    """Create a compile-only workbook with dangerous defined names removed.

    1) Strip hard-broken #REF! names (e.g. TriangleRange =#REF!#REF!) so pycel
       cannot propagate the error into the objective.
    2) For duplicate same-name definitions (e.g. workbook-scoped + sheet-scoped
       `truncation`), drop only the copy whose single-cell target is
       blank/unsupported when a sibling points at a usable cell. Valid scoped
       names are NOT blindly dropped — Excel scope rules still matter.
    """
    from openpyxl import load_workbook

    src_xlsx = Path(src_xlsx)
    try:
        wb = load_workbook(src_xlsx, data_only=False, keep_links=False)
    except Exception as e:
        if log_fn:
            log_fn(f"Defined-name sanitize skipped: could not open workbook ({e})")
        return src_xlsx

    removed = []

    # 1) hard-broken #REF! names
    for container, d, name, target, sheet in list(_iter_defined_name_entries(wb)):
        if "#REF!" in str(target).upper():
            if _remove_defined_name_from_container(container, d, name):
                removed.append(f"{sheet or 'WORKBOOK'}!{name}->{target}")

    # 2) duplicate names: drop only blank/unsupported copies when a usable
    #    sibling exists
    entries = list(_iter_defined_name_entries(wb))
    groups = {}
    for entry in entries:
        groups.setdefault(entry[2].casefold(), []).append(entry)
    for group in groups.values():
        if len(group) < 2:
            continue
        scored = [
            (_defined_name_single_cell_value(wb, target) is not None,
             container, d, name, target, sheet)
            for container, d, name, target, sheet in group
        ]
        usable_flags = [u for u, *_ in scored]
        if not any(usable_flags) or all(usable_flags):
            continue
        for usable, container, d, name, target, sheet in scored:
            if usable:
                continue
            if _remove_defined_name_from_container(container, d, name):
                removed.append(f"duplicate blank {sheet or 'WORKBOOK'}!{name}->{target}")

    if not removed:
        if log_fn:
            log_fn("Defined-name sanitize: nothing to remove")
        return src_xlsx

    out = src_xlsx.with_name("_fastsolver_pycel_names_sanitized.xlsx")
    try:
        wb.save(out)
        if log_fn:
            shown = "; ".join(removed[:8])
            extra = "" if len(removed) <= 8 else f"; +{len(removed)-8} more"
            log_fn(f"Defined-name sanitize: removed {len(removed)} broken names ({shown}{extra})")
        return out
    except Exception as e:
        if log_fn:
            log_fn(f"Defined-name sanitize save failed ({e}); using unsanitized workbook")
        return src_xlsx


# ---------------------------------------------------------------------------
# Pycel numeric guard
# ---------------------------------------------------------------------------
def _install_pycel_numeric_guard():
    """Make pycel return Excel-style error values for numeric blow-ups.

    pycel 1.0b30 evaluates formulas as Python expressions. For log-log /
    Pareto fits, optimizers may probe values that make ``base ** shape``
    enormous or invalid. Excel turns those into #NUM!/#DIV/0! and continues;
    pycel can raise OverflowError before _safe_eval gets a value back.
    Guarding the operator table keeps bad probes as penalty scores.
    """
    try:
        import pycel.excelutil as _eu
        ops = getattr(_eu, "PYTHON_AST_OPERATORS", None)
        if not isinstance(ops, dict):
            return

        for key, fn in list(ops.items()):
            if not callable(fn) or getattr(fn, "_fastsolver_guarded", False):
                continue

            def _guarded_operator(*args, _fn=fn):
                """Guard pycel numeric operators without breaking unary ops.

                pycel uses the same operator table for binary (Add/Sub/Mult/
                Div/Pow) and unary (USub/UAdd) operators. Accept *args and
                forward exactly what pycel supplied so ``x ** -shape`` works.
                """
                try:
                    return _fn(*args)
                except ZeroDivisionError:
                    return "#DIV/0!"
                except OverflowError:
                    return "#NUM!"
                except ValueError:
                    return "#NUM!"
                except TypeError:
                    for a in args:
                        if isinstance(a, str) and a.startswith("#"):
                            return a
                    return "#VALUE!"

            _guarded_operator._fastsolver_guarded = True
            ops[key] = _guarded_operator
    except Exception:
        pass


_install_pycel_numeric_guard()

from pycel import ExcelCompiler
from scipy.optimize import minimize, differential_evolution
from openpyxl import load_workbook


# ===========================================================================
# Helpers — simplex projection, bounds, etc.
# ===========================================================================

def project_simplex_group(x, idxs, lo=None, hi=None):
    """In place: project x[idxs] onto {sum == 1, lo <= x <= hi}.

    The previous version projected onto sum==1 only, ignoring the per-
    variable bounds. With bounded weights (e.g. A1,B1 in [0.4,0.6], Sum=1)
    that produced two failures:
      1. a single-coordinate perturbation (how coordinate descent moves)
         was almost entirely cancelled by the renormalisation, so CD saw
         "no change", shrank the step to min_step and stalled;
      2. the renormalised values could leave [lo,hi] with no re-clip.

    Projecting onto the bounded simplex fixes both: the feasible set is
    respected, and a one-variable nudge maps to a real, non-cancelled move.
    """
    sub = np.asarray(x[idxs], float)
    k = len(sub)

    if lo is None or hi is None:
        sub = np.maximum(sub, 0.0)
        s = sub.sum()
        x[idxs] = (sub / s) if s > 0 else (np.ones(k) / k)
        return

    glo = np.asarray(lo, float)[idxs]
    ghi = np.asarray(hi, float)[idxs]

    # Feasible only if sum of bounds straddles 1; otherwise fall back to the
    # closest corner so we never emit NaN/inf.
    if glo.sum() > 1.0 + 1e-12:
        x[idxs] = glo.copy()
        return
    if ghi.sum() < 1.0 - 1e-12:
        x[idxs] = ghi.copy()
        return

    # Bisection on the Lagrange multiplier for sum(clip(sub - t)) == 1.
    sub = np.clip(sub, glo, ghi)
    lo_t = float(np.min(sub - ghi))
    hi_t = float(np.max(sub - glo))
    for _ in range(60):
        t = 0.5 * (lo_t + hi_t)
        proj = np.clip(sub - t, glo, ghi)
        if proj.sum() > 1.0:
            lo_t = t
        else:
            hi_t = t
    x[idxs] = np.clip(sub - 0.5 * (lo_t + hi_t), glo, ghi)


def collect_simplex_groups(variables):
    """Return list of index-arrays for variables sharing rule 'Sum=1'."""
    groups = {}
    for i, v in enumerate(variables):
        if v["rule"] == "Sum=1":
            groups.setdefault(v["group"], []).append(i)
    return [np.array(idxs, dtype=int) for idxs in groups.values()]


def apply_projection(x, simplex_groups, lo, hi):
    x = np.clip(x, lo, hi)
    for idxs in simplex_groups:
        project_simplex_group(x, idxs, lo, hi)
    # Re-clip: a sum=1 projection can nudge a value back outside its box,
    # and downstream code (and the user's stated bounds) assume feasibility.
    x = np.clip(x, lo, hi)
    return x


# ===========================================================================
# Excel evaluation via pycel — score function builder
# ===========================================================================

def _split_addr(ref):
    """Split 'SheetName!$A$1' or "'Sheet 1'!$A$1" → (sheet, cell). Tolerates missing leading quote."""
    if "!" not in ref:
        return None, ref
    sheet, cell = ref.split("!", 1)
    sheet = sheet.strip().strip("'").strip()
    cell = cell.replace("$", "")
    return sheet, cell


def _pycel_addr(ref):
    """Convert VBA-exported ref → pycel address format SheetName!A1."""
    sheet, cell = _split_addr(ref)
    if sheet is None:
        return cell
    if " " in sheet:
        return f"'{sheet}'!{cell}"
    return f"{sheet}!{cell}"


def build_score_fn(excel, config, variables, simplex_groups, lo, hi):
    name_maps = getattr(excel, "_fastsolver_name_maps", None)
    default_sheet = _split_addr(variables[0]["ref"])[0] if variables else None

    def _ref_addr(ref, current_sheet=None):
        resolved = _resolve_defined_name_reference(ref, current_sheet or default_sheet, name_maps)
        return _pycel_addr(resolved or ref)

    var_addrs = [_ref_addr(v["ref"], _split_addr(v["ref"])[0] or default_sheet) for v in variables]
    obj_specs = [o for o in config["objectives"] if o["active"]]
    con_specs = [c for c in config["constraints"] if c["active"]]
    alpha = float(config.get("outlier_alpha", 0.7))
    hard_pen = float(config.get("hard_penalty", 1e6))

    eval_count = [0]

    def write_vars(x):
        for addr, val in zip(var_addrs, x):
            try:
                # pycel 1.0b30 skips invalidation when the written value is
                # unchanged. Clear the input cell first so downstream formulas
                # are reset even when optimizers re-probe the same point.
                try:
                    c = excel.cell_map.get(addr)
                    if c is not None:
                        c.value = None
                except Exception:
                    pass
                excel.set_value(addr, float(val))
            except Exception:
                pass

    # ---- Memoization -----------------------------------------------------
    _memo = {}
    _MEMO_DECIMALS = 10

    def _memo_key(x):
        return tuple(round(float(v), _MEMO_DECIMALS) for v in x)

    formula_cache = {}

    def eval_ref(spec_str, is_rhs=False):
        s = spec_str.strip()
        host_sheet = default_sheet or "Sheet1"
        if not s.startswith("="):
            return _safe_eval(excel, _ref_addr(s, host_sheet))

        try:
            s_for_pycel = _rewrite_formula_defined_names(s, host_sheet, name_maps)
        except Exception:
            s_for_pycel = s

        if s_for_pycel in formula_cache:
            tmp_addr = formula_cache[s_for_pycel]
        else:
            tmp_addr = f"{host_sheet}!ZZ{9000 + len(formula_cache)}"
            try:
                excel.set_value(tmp_addr, s_for_pycel)
            except Exception:
                return 1e12
            formula_cache[s_for_pycel] = tmp_addr

        try:
            cell = excel.cell_map.get(tmp_addr)
            if cell is not None:
                cell.value = None
            v = excel.evaluate(tmp_addr)
            if isinstance(v, (list, tuple)):
                flat = []
                for r in v:
                    if isinstance(r, (list, tuple)): flat.extend(r)
                    else: flat.append(r)
                return float(sum(x for x in flat if isinstance(x, (int, float))))
            return float(v or 0)
        except Exception:
            return 1e12

    def _score_raw(x):
        eval_count[0] += 1
        x = apply_projection(x, simplex_groups, lo, hi)
        write_vars(x)

        any_obj = False
        total_weighted = 0.0
        max_signed = -1e30
        for o in obj_specs:
            val = eval_ref(o["ref"])
            goal = (o["goal"] or "Min").lower()
            if goal == "min":
                s_val = val
            elif goal == "max":
                s_val = -val
            elif goal == "target":
                s_val = abs(val - float(o.get("target", 0)))
            else:
                s_val = val
            w = float(o.get("weight", 1.0))
            total_weighted += w * s_val
            if w * s_val > max_signed: max_signed = w * s_val
            any_obj = True

        obj_part = (alpha * max_signed + (1 - alpha) * total_weighted) if any_obj else 0.0

        penalty = 0.0
        for c in con_specs:
            lhs = eval_ref(c["lhs"])
            rhs_str = c["rhs"]
            try:
                rhs = float(rhs_str)
            except (TypeError, ValueError):
                rhs = eval_ref(rhs_str, True)
            diff = lhs - rhs
            op = c.get("op", "=")
            if op == "=":
                viol = abs(diff)
            elif op in ("<=", "<"):
                viol = max(0.0, diff)
            elif op in (">=", ">"):
                viol = max(0.0, -diff)
            else:
                viol = 0.0
            pw = hard_pen if c.get("type", "Hard").lower() == "hard" else float(c.get("penalty", 1000))
            penalty += pw * viol * viol

        try:
            out = float(obj_part + penalty)
            if out != out or abs(out) == float('inf'):
                return 1e12
            return out
        except Exception:
            return 1e12

    def score(x):
        try:
            k = _memo_key(x)
            cached = _memo.get(k)
            if cached is not None:
                return cached
            v = _score_raw(x)
            try:
                fv = float(v)
            except Exception:
                return 1e12
            if fv != fv or abs(fv) == float('inf'):
                return 1e12
            # Do not memoize sentinel/error scores. A transient pycel #DIV/0!,
            # #VALUE!, #NUM!/overflow, or evaluation exception should not
            # poison later runs that revisit the same rounded point after the
            # graph has been reset.
            if fv < 1e12:
                _memo[k] = fv
            return fv
        except Exception:
            return 1e12

    score.eval_count = eval_count
    score.var_addrs = var_addrs
    score.write_vars = write_vars
    score.clear_memo = _memo.clear
    return score


def _safe_eval(excel, addr):
    try:
        # pycel 1.0b30 can leave formula cells with stale error values after
        # a bad probe, especially through branching dependency graphs. Clear
        # the requested formula cell before reading so evaluate() recomputes it.
        try:
            c = excel.cell_map.get(addr)
            if c is not None and getattr(c, "formula", None):
                c.value = None
        except Exception:
            pass

        v = excel.evaluate(addr)
        if v is None:
            return 1e12
        if isinstance(v, str) and v.startswith("#"):
            return 1e12
        fv = float(v)
        if fv != fv or abs(fv) == float('inf'):
            return 1e12
        return fv
    except Exception:
        return 1e12


# ===========================================================================
# Coordinate Descent — Anti-Zigzag + Adaptive Step
# ===========================================================================

def coordinate_descent_azas(score, x0, lo, hi, simplex_groups,
                            max_evals=500, init_step=0.1,
                            min_step=1e-7, tol=1e-9, log_fn=None):
    """Robust bounded coordinate descent with anti-zigzag pattern moves.

    Intentionally does NOT use scipy Powell. Powell's line search can probe
    far outside the valid basin; for log-log/Pareto sheets a single large
    probe produces #NUM!/overflow and the line search gives up. This version
    is conservative: try +/- one variable at a time inside bounds; shrink
    steps when both directions are bad; grow a step only after it improves;
    after each improving pass, try one pattern move in the net direction.
    Bad probes score 1e12 and are rejected.
    """
    def log(s):
        if log_fn:
            log_fn(s)

    n = len(x0)
    x_best = apply_projection(np.asarray(x0, float).copy(), simplex_groups, lo, hi)
    f_best = float(score(x_best))

    if n == 0:
        return x_best, f_best

    lo_a = np.asarray(lo, float)
    hi_a = np.asarray(hi, float)
    span = hi_a - lo_a
    finite_span = np.where(np.isfinite(span) & (span > 0), span, 1.0)

    # Coordinate descent is very sensitive to step scale. A raw 10% of range
    # is fine for a small bounded weight like [0.4, 0.6], but disastrous for
    # positive scale parameters such as [500, 500_000_000]: the first probe is
    # ~50 million units away from the incumbent and usually lands in an Excel
    # #NUM!/overflow/flat-penalty region. Detect those wide positive ranges and
    # move them multiplicatively in log10 space instead.
    # Compute range ratios without evaluating hi/lo where lo is zero.
    # np.where still evaluates both branches, so use np.divide(..., where=...)
    # to avoid RuntimeWarning on shape variables whose lower bound is 0.
    range_ratio = np.ones(n, dtype=float)
    np.divide(hi_a, lo_a, out=range_ratio,
              where=(lo_a > 0.0) & (hi_a > 0.0))

    wide_positive = (
        np.isfinite(lo_a) & np.isfinite(hi_a)
        & (lo_a > 0.0) & (hi_a > 0.0)
        & (range_ratio >= 1e4)
    )

    step = np.maximum(init_step * finite_span, min_step)
    min_step_vec = np.full(n, float(min_step))
    max_step_vec = finite_span.copy()

    # For wide-positive variables, step[i] is a log10 delta.  log10(1.10)
    # means the first +/- trial is roughly +/-10% multiplicative, e.g.
    # 522_000 -> 574_000 or 475_000, instead of 522_000 +/- 50_000_000.
    if np.any(wide_positive):
        step[wide_positive] = max(float(np.log10(1.0 + init_step)), 1e-5)
        min_step_vec[wide_positive] = 1e-6
        max_step_vec[wide_positive] = 0.5  # at most ~3.16x per coordinate try
        log(f"  CD-AZAS: using multiplicative log steps on "
            f"{int(np.sum(wide_positive))}/{n} wide-positive variables")

    no_improve_limit = 12

    start_evals = score.eval_count[0]
    passes = 0
    no_improve_passes = 0

    log(f"  CD-AZAS start: f={f_best:.8g}, budget={max_evals}")

    while score.eval_count[0] - start_evals < max_evals:
        passes += 1
        x_pass_start = x_best.copy()
        f_pass_start = f_best
        improved_this_pass = False

        order = np.argsort(-step)
        for i in order:
            if score.eval_count[0] - start_evals >= max_evals:
                break
            if step[i] <= min_step_vec[i]:
                continue

            base = x_best.copy()
            trials = []
            for direction in (1.0, -1.0):
                z = base.copy()
                if wide_positive[i]:
                    # Multiplicative/log-space move for scale parameters.
                    # Clip the exponent to avoid accidental extreme probes.
                    exponent = float(np.clip(direction * step[i], -0.5, 0.5))
                    z[i] = max(base[i], lo_a[i]) * (10.0 ** exponent)
                else:
                    z[i] = z[i] + direction * step[i]
                z = apply_projection(z, simplex_groups, lo, hi)
                trials.append(z)

            best_trial_x = None
            best_trial_f = f_best
            for z in trials:
                if score.eval_count[0] - start_evals >= max_evals:
                    break
                fz = float(score(z))
                if fz < best_trial_f - tol * (1.0 + abs(best_trial_f)):
                    best_trial_x, best_trial_f = z, fz

            if best_trial_x is not None:
                x_best, f_best = best_trial_x, best_trial_f
                improved_this_pass = True
                step[i] = min(step[i] * 1.35, max_step_vec[i])
            else:
                step[i] *= 0.5

        if improved_this_pass and score.eval_count[0] - start_evals < max_evals:
            direction = x_best - x_pass_start
            if np.any(np.abs(direction) > 0):
                z = x_best + direction

                # Pattern moves should match the coordinate move geometry.
                # For log-stepped scale variables, repeat the multiplicative
                # ratio from the pass instead of using a raw additive delta.
                for j in np.where(wide_positive)[0]:
                    if x_pass_start[j] > 0.0 and x_best[j] > 0.0:
                        ratio = np.clip(x_best[j] / x_pass_start[j],
                                        10.0 ** -0.5, 10.0 ** 0.5)
                        z[j] = x_best[j] * ratio

                z = apply_projection(z, simplex_groups, lo, hi)
                fz = float(score(z))
                if fz < f_best - tol * (1.0 + abs(f_best)):
                    x_best, f_best = z, fz

        if improved_this_pass:
            no_improve_passes = 0
            log(f"  CD-AZAS pass {passes}: f={f_best:.8g}, "
                f"evals={score.eval_count[0] - start_evals}")
        else:
            no_improve_passes += 1
            step *= 0.5
            if no_improve_passes == 1 or no_improve_passes % 3 == 0:
                try:
                    max_step_ratio = float(np.nanmax(step / np.maximum(min_step_vec, 1e-300)))
                except Exception:
                    max_step_ratio = float('nan')
                log(f"  CD-AZAS no-improve pass {no_improve_passes}/{no_improve_limit}; "
                    f"shrinking steps (max step/min = {max_step_ratio:.3g})")
            if no_improve_passes >= no_improve_limit:
                break

        if np.nanmax(step / np.maximum(min_step_vec, 1e-300)) <= 1.0:
            break

        # Important: do NOT stop after the first no-improvement pass.
        # The whole point of adaptive CD is to reject too-large moves, shrink,
        # and retry. The previous check below made no_improve_limit ineffective
        # and caused logs like: done after evals=9, passes=1.
        #
        #     if abs(f_pass_start - f_best) <= tol * (...) and not improved_this_pass:
        #         break

    log(f"  CD-AZAS done: f={f_best:.8g}, "
        f"evals={score.eval_count[0] - start_evals}, passes={passes}")
    return x_best, f_best


# ===========================================================================
# SciPy Powell direction-set minimiser  (used by portfolio-powell)
# ===========================================================================
# Powell intrinsic strengths that make it a good portfolio leg:
#   - Anti-zigzag: maintains conjugate direction set, replaces worst axis
#     each iteration with the net-progress direction (Nocedal Ch.9.4)
#   - Adaptive step: internal Brent line search auto-expands/contracts
#   - score() already maps bad probes (NaN/inf/#NUM!) to 1e12
#   - projected() wrapper re-clips/re-simplexes every probe
#   - SciPy Powell gets explicit bounds for the hardened bounded variant
#   - unexpected failure falls back to CD-AZAS (never worse than not picking it)
# ===========================================================================

def scipy_powell(score, x0, lo, hi, simplex_groups,
                 max_evals=500, tol=1e-9, log_fn=None):
    """Powell direction-set minimisation through the memoised pycel score."""
    def log(s):
        if log_fn:
            log_fn(s)

    n = len(x0)
    x0p = apply_projection(np.asarray(x0, float).copy(),
                           simplex_groups, lo, hi)
    f0 = float(score(x0p))

    if n == 0:
        return x0p, f0

    def projected(z):
        return score(apply_projection(np.asarray(z, float).copy(),
                                      simplex_groups, lo, hi))

    log(f"  Powell start: f={f0:.8g}, budget={max_evals}")

    try:
        res = minimize(projected, x0p, method="Powell",
                       bounds=list(zip(lo, hi)),
                       options={"maxfev": int(max_evals),
                                "xtol": 1e-8,
                                "ftol": 1e-10})
        x_final = apply_projection(np.asarray(res.x, float).copy(),
                                   simplex_groups, lo, hi)
        f_final = float(score(x_final))

        if f_final > f0:
            log(f"  Powell: polished point worse ({f_final:.6g} > "
                f"{f0:.6g}); keeping start point")
            x_final, f_final = x0p, f0

        log(f"  Powell done: f={f_final:.8g}, "
            f"evals={score.eval_count[0]}, "
            f"converged={getattr(res, 'success', '?')}")
        return x_final, f_final

    except Exception as e:
        log(f"  Powell raised ({e}); falling back to CD-AZAS")
        return coordinate_descent_azas(
            score, x0p, lo, hi, simplex_groups,
            max_evals=max_evals, tol=tol, log_fn=log_fn)


# ===========================================================================
# Method dispatch
# ===========================================================================

def _method_key(method):
    """Normalize UI/VBA method labels so spelling/spacing changes do not
    silently route to the wrong solver."""
    s = str(method or "auto").strip().lower()
    for ch in "_–—/()":
        s = s.replace(ch, " ")
    s = s.replace("-", " ")
    s = " ".join(s.split())
    aliases = {
        "auto": "auto",
        "slsqp": "slsqp",
        "slsqp python": "slsqp",
        "l bfgs b": "l-bfgs-b",
        "lbfgs": "l-bfgs-b",
        "homotopy autodiff": "l-bfgs-b",
        "homotopy": "l-bfgs-b",
        "differential evolution": "de",
        "de": "de",
        "de multi restart": "de-multi-restart",
        "differential evolution multi restart": "de-multi-restart",
        "de 100": "de-multi-restart",
        "de100": "de-multi-restart",
        "portfolio": "portfolio",
        "portfolio all methods": "portfolio",
        "best of all": "portfolio",
        "round robin": "portfolio",
        "portfolio cd": "portfolio-cd",
        "portfolio powell": "portfolio-powell",
        "powell": "powell",
        "coordinate descent": "cd-azas",
        "cordinate descent": "cd-azas",
        "coordinate decent": "cd-azas",
        "cordinate decents": "cd-azas",
        "coordinate descents": "cd-azas",
        "coordinate descent anti zigzag": "cd-azas",
        "coordinate descent anti zigzag adaptive step": "cd-azas",
        "cd": "cd-azas",
        "cd azas": "cd-azas",
        "cd-azas": "cd-azas",
        "coordinate descent azas": "cd-azas",
        "coordinate_descent_azas": "cd-azas",
    }
    return aliases.get(s, s)


def run_method(method, score, x0, bounds, simplex_groups, lo, hi, max_iter,
               de_seed=None, de_seed_with_x0=False):
    method = _method_key(method)
    n = len(x0)

    def projected(x):
        return score(apply_projection(np.asarray(x, float).copy(), simplex_groups, lo, hi))

    if method == "auto":
        method = "l-bfgs-b" if n > 50 else "slsqp"

    if method == "portfolio":
        return portfolio_solve(score, x0, bounds, simplex_groups, lo, hi,
                               max_iter, log_fn=getattr(score, "log_fn", None))

    if method == "cd-azas":
        x_final, f_final = coordinate_descent_azas(
            score, x0, lo, hi, simplex_groups,
            max_evals=max_iter, init_step=0.1,
            min_step=1e-7, tol=1e-9,
            log_fn=getattr(score, "log_fn", None))
        return x_final, f_final

    if method == "slsqp":
        res = minimize(projected, x0, method="SLSQP",
                       bounds=list(zip(lo, hi)),
                       options={"maxiter": max_iter, "ftol": 1e-9})
        x_final = apply_projection(res.x, simplex_groups, lo, hi)
        return x_final, res.fun

    if method == "l-bfgs-b":
        res = minimize(projected, x0, method="L-BFGS-B",
                       bounds=list(zip(lo, hi)),
                       options={"maxiter": max_iter, "ftol": 1e-9})
        x_final = apply_projection(res.x, simplex_groups, lo, hi)
        return x_final, res.fun

    if method == "de":
        if de_seed is None:
            import secrets
            this_seed = secrets.randbelow(2**31 - 1)
        else:
            this_seed = int(de_seed) % (2**31 - 1)

        log_fn = getattr(score, "log_fn", None)
        lo_a = np.asarray(lo, float)
        hi_a = np.asarray(hi, float)

        # ---- Variable rescaling -----------------------------------------
        # DE uses a single differential weight across all dimensions, so a
        # variable on [500, 5e8] sitting next to one on [0, 5] keeps the
        # population variance permanently huge -> std(pop) <= tol is never
        # reached -> DE always exits "Maximum iterations exceeded". Optimise
        # in a normalised space instead: log10 for wide positive ranges,
        # linear unit-scaling otherwise. DE sees a well-conditioned [0,1]^n
        # cube; we map back to real units before every score() call.
        WIDE = 1e4  # hi/lo ratio above which we switch to log scaling
        use_log = np.zeros(n, dtype=bool)
        for i in range(n):
            if lo_a[i] > 0 and hi_a[i] > 0 and (hi_a[i] / lo_a[i]) >= WIDE:
                use_log[i] = True

        llo = np.where(use_log, np.log10(np.where(lo_a > 0, lo_a, 1.0)), lo_a)
        lhi = np.where(use_log, np.log10(np.where(hi_a > 0, hi_a, 1.0)), hi_a)
        lspan = np.where(lhi > llo, lhi - llo, 1.0)

        def _to_real(u):
            u = np.asarray(u, float)
            raw = llo + np.clip(u, 0.0, 1.0) * lspan
            real = np.where(use_log, np.power(10.0, raw), raw)
            return np.clip(real, lo_a, hi_a)

        def _scaled_obj(u):
            return projected(_to_real(u))

        def _to_unit(xr):
            xr = np.clip(np.asarray(xr, float), lo_a, hi_a)
            lin = np.where(use_log,
                           np.log10(np.where(xr > 0, xr, 1.0)),
                           xr)
            return np.clip((lin - llo) / lspan, 0.0, 1.0)

        # ---- Budget ------------------------------------------------------
        # SciPy DE's maxiter is GENERATIONS, not evaluations. One generation
        # costs roughly popsize*n workbook evaluations. In portfolio mode the
        # caller may pass a large eval budget; letting DE convert that into
        # hundreds of generations can run for minutes with no progress-file
        # update, which looks like a crash to the VBA watchdog. Treat max_iter
        # as an evaluation budget and size generations from popsize*n.
        popsize = 10 if n <= 8 else 8
        evals_per_generation = max(1, popsize * max(1, n))
        gens = int(max(8, min(80, max_iter // evals_per_generation)))

        prog_dir = getattr(score, "run_dir", None)
        de_last_progress = {"gen": 0}

        def _de_callback(xk, convergence=None):
            # Called once per generation. Keep progress.txt fresh so the Excel
            # side does not assume the solver hung while DE is still working.
            de_last_progress["gen"] += 1
            if prog_dir and (de_last_progress["gen"] == 1 or
                             de_last_progress["gen"] % 3 == 0 or
                             de_last_progress["gen"] >= gens):
                try:
                    incumbent = score(_to_real(xk))
                except Exception:
                    incumbent = float("nan")
                _write_progress(
                    prog_dir,
                    f"DE generation {de_last_progress['gen']}/{gens} | "
                    f"evals={score.eval_count[0]} | best~{incumbent:.6g}"
                )
            return False

        de_kwargs = dict(
            bounds=[(0.0, 1.0)] * n,
            maxiter=gens,
            popsize=popsize,
            tol=1e-7,
            mutation=(0.5, 1.2),
            recombination=0.8,
            seed=this_seed,
            polish=False,
            init="sobol",
            callback=_de_callback,
            updating="immediate",
            workers=1,
        )

        if de_seed_with_x0:
            try:
                rng = np.random.default_rng(this_seed)
                pop = rng.random((popsize, n))
                pop[0] = _to_unit(np.asarray(x0, float))
                de_kwargs["init"] = pop
            except Exception:
                pass

        res = differential_evolution(_scaled_obj, **de_kwargs)

        x_final = apply_projection(_to_real(res.x), simplex_groups, lo, hi)
        f_final = score(x_final)

        # DE stopping on maxiter is NORMAL for a global, derivative-free
        # search over a flat/multi-modal landscape -- it is not an error.
        # Compare against the starting point and keep whichever is better,
        # rather than discarding a large improvement just because
        # res.success is False.
        try:
            x_start = apply_projection(np.asarray(x0, float).copy(),
                                       simplex_groups, lo, hi)
            f_start = score(x_start)
        except Exception:
            f_start = float("inf")
            x_start = x_final

        if f_start < f_final:
            x_final, f_final = x_start, f_start

        if log_fn:
            conv = "converged" if res.success else \
                   "stopped on maxiter (normal for DE)"
            log_fn(f"  DE: {conv}; f={f_final:.8g} "
                   f"(start {f_start:.6g}), {gens} gens, "
                   f"{score.eval_count[0]} evals, "
                   f"log-scaled dims={int(use_log.sum())}/{n}")

        return x_final, f_final


    if method == "de-multi-restart":
        # Run DE up to N times with different seeds, keeping the best. The
        # incumbent is injected into every restart's population so a run can
        # never lose the best basin already found. Early exits: patience
        # (consecutive no-improvement restarts) and the caller's eval budget.
        N_RESTARTS = 2
        patience = 1

        best_x = apply_projection(np.asarray(x0, float).copy(),
                                  simplex_groups, lo, hi)
        best_f = score(best_x)
        log_fn = getattr(score, "log_fn", None)
        prog_dir = getattr(score, "run_dir", None)

        no_improve = 0
        start_evals = score.eval_count[0]
        for k in range(1, N_RESTARTS + 1):
            if score.eval_count[0] - start_evals >= max_iter:
                if log_fn:
                    log_fn(f"  DE-MultiRestart: eval budget reached at "
                           f"restart {k}; returning best f={best_f:.8g}")
                break

            this_seed = (k * 2654435761) % (2**31 - 1)
            try:
                popn = max(5, 15)
                rng = np.random.default_rng(this_seed)
                pop = lo + rng.random((popn, n)) * (hi - lo)
                pop[0] = np.clip(best_x, lo, hi)
                init = pop
            except Exception:
                init = "latinhypercube"

            r = differential_evolution(
                projected, bounds=list(zip(lo, hi)),
                maxiter=max(20, (max_iter // 10)),
                tol=1e-7, seed=this_seed, polish=True, init=init)

            rx = apply_projection(r.x, simplex_groups, lo, hi)
            rf = score(rx)
            if rf < best_f - 1e-12 * (1.0 + abs(best_f)):
                best_x, best_f = rx, rf
                no_improve = 0
                if log_fn:
                    log_fn(f"  DE-MultiRestart {k}/{N_RESTARTS}: "
                           f"new best f={best_f:.8g}")
                if prog_dir:
                    _write_progress(prog_dir,
                                    f"DE restart {k}/{N_RESTARTS} | "
                                    f"NEW BEST f={best_f:.6g}")
            else:
                no_improve += 1
                if prog_dir:
                    _write_progress(prog_dir,
                                    f"DE restart {k}/{N_RESTARTS} | "
                                    f"no gain ({no_improve}/{patience}) | "
                                    f"best f={best_f:.6g}")

            if no_improve >= patience:
                if log_fn:
                    log_fn(f"  DE-MultiRestart: {patience} restarts with no "
                           f"improvement (stopped at {k}); f={best_f:.8g}")
                break

        return best_x, best_f

    if method == "powell":
        x_final, f_final = scipy_powell(
            score, x0, lo, hi, simplex_groups,
            max_evals=max_iter, tol=1e-9,
            log_fn=getattr(score, "log_fn", None))
        return x_final, f_final

    # Default
    res = minimize(projected, x0, method="L-BFGS-B",
                   bounds=list(zip(lo, hi)),
                   options={"maxiter": max_iter, "ftol": 1e-9})
    x_final = apply_projection(res.x, simplex_groups, lo, hi)
    return x_final, res.fun


# ===========================================================================
# Portfolio meta-solver  (round-robin, interpretation A)
# ===========================================================================
def portfolio_solve(score, x0, bounds, simplex_groups, lo, hi,
                    max_iter, log_fn=None,
                    order=("slsqp", "l-bfgs-b",
                           "coordinate descent anti-zigzag adaptive step",
                           "differential evolution"),
                    rel_tol=1e-9, max_rounds=100,
                    no_change_limit=2):
    """Run a portfolio of methods, chaining improvements.

    For each method: run from the current best; if it improved, re-run the
    SAME method from the new point until it stops improving; then advance.
    After a full pass, if anything improved, cycle again. Stop after
    no_change_limit no-improvement passes, max_rounds, or eval budget.

    Every method shares the memoized `score`, so re-running from an explored
    point is mostly free cache hits.
    """
    def log(s):
        if log_fn:
            log_fn(s)

    _run_dir = getattr(score, "run_dir", None)

    def progress(s):
        if _run_dir:
            _write_progress(_run_dir, s)

    best_x = apply_projection(np.asarray(x0, float).copy(),
                              simplex_groups, lo, hi)
    best_f = score(best_x)
    log(f"Portfolio start: f={best_f:.8g}")
    progress(f"Starting. Initial score = {best_f:.6g}")

    def improved(new_f, old_f):
        return new_f < old_f - rel_tol * (1.0 + abs(old_f))

    total_start_evals = score.eval_count[0]
    no_change_rounds = 0

    # Budget each portfolio method. DE is intentionally capped lower because
    # one DE generation costs popsize*n pycel workbook evaluations and can be
    # much slower than local methods.
    per_call_budget = max(50, max_iter // 12)
    de_call_budget = max(300, min(2500, max_iter // 60))

    for rnd in range(1, max_rounds + 1):
        round_improved = False
        log(f"--- Portfolio round {rnd} ---")
        progress(f"Round {rnd} starting (best so far = {best_f:.6g})")

        for method in order:
            method_iters = 0
            while True:
                if score.eval_count[0] - total_start_evals >= max_iter:
                    log(f"  [{method}] global eval budget reached; "
                        f"stopping portfolio")
                    progress(f"Eval budget reached - finishing. "
                             f"Best = {best_f:.6g}")
                    return best_x, best_f

                evals_now = score.eval_count[0] - total_start_evals
                progress(f"Round {rnd} | {method} | attempt {method_iters + 1} "
                         f"| {evals_now} evals | best = {best_f:.6g}")

                if method == "de":
                    de_seed = 1000 * rnd + 7 * (method_iters + 1) + 13
                    x_try, f_try = run_method(method, score, best_x.copy(),
                                              bounds, simplex_groups, lo, hi,
                                              de_call_budget,
                                              de_seed=de_seed,
                                              de_seed_with_x0=True)
                else:
                    x_try, f_try = run_method(method, score, best_x.copy(),
                                              bounds, simplex_groups, lo, hi,
                                              per_call_budget)

                method_iters += 1

                if improved(f_try, best_f):
                    delta = best_f - f_try
                    best_x, best_f = x_try, f_try
                    round_improved = True
                    log(f"  [{method}] improved -> f={best_f:.8g} "
                        f"(-{delta:.3g}); re-running same method")
                    progress(f"  -> {method} improved! "
                             f"new best = {best_f:.6g} (re-running it)")
                else:
                    if method_iters == 1:
                        log(f"  [{method}] no improvement (f={f_try:.8g})")
                    else:
                        log(f"  [{method}] exhausted after {method_iters} "
                            f"runs; best f={best_f:.8g}")
                    progress(f"  -> {method} done (no further gain)")
                    break

        if round_improved:
            no_change_rounds = 0
        else:
            no_change_rounds += 1
            log(f"Portfolio no-change round {no_change_rounds}/"
                f"{no_change_limit} (round {rnd}), f={best_f:.8g}")
            progress(f"No change {no_change_rounds}/{no_change_limit} "
                     f"| best = {best_f:.6g}")

            if no_change_rounds >= no_change_limit:
                log(f"Portfolio converged: {no_change_limit} consecutive "
                    f"full rounds with no improvement, f={best_f:.8g}")
                progress(f"Converged after {no_change_limit} no-change rounds. "
                         f"Best = {best_f:.6g}")
                break
    else:
        log(f"Portfolio hit max_rounds={max_rounds}; f={best_f:.8g}")

    return best_x, best_f


# ===========================================================================
# Multi-start wrapper
# ===========================================================================

def multi_start_solve(method, score, x0, bounds, simplex_groups, lo, hi,
                      max_iter, n_starts=3):
    best_x, best_f = None, np.inf
    rng = np.random.default_rng(42)
    n = len(x0)
    log_fn = getattr(score, "log_fn", None)
    for s in range(n_starts):
        if s == 0:
            x_init = x0.copy()
        else:
            x_init = np.clip(x0 + rng.normal(0, 0.2, n), lo, hi)
            x_init = apply_projection(x_init, simplex_groups, lo, hi)
        if log_fn:
            log_fn(f"  Multi-start {s + 1}/{n_starts}: method={_method_key(method)}")
        x, f = run_method(method, score, x_init, bounds, simplex_groups,
                           lo, hi, max_iter)
        if f < best_f:
            best_f, best_x = f, x
    return best_x, best_f


# ===========================================================================
# Main driver
# ===========================================================================

def _compute_cache_key(src_xlsx, variables, log_fn):
    """Build a cache key that ignores ONLY the variable cells.

    The compiled pycel graph depends on formulas and the constants formulas
    read — not on variable cells, which the solver overwrites anyway. Hashing
    with variable cells excluded means the solver's own output never busts
    the cache, while any user edit still forces a correct recompile.
    """
    import hashlib
    from openpyxl import load_workbook

    try:
        wb = load_workbook(src_xlsx, data_only=False)
    except Exception as e:
        log_fn(f"Cache key: workbook introspection failed ({e}); using raw hash")
        return hashlib.sha1(Path(src_xlsx).read_bytes()).hexdigest()[:16]

    var_cells = set()
    for v in variables:
        ref = v.get("ref", "")
        if "!" not in ref:
            continue
        sheet, cell = ref.split("!", 1)
        sheet = sheet.strip().strip("'").strip()
        cell = cell.replace("$", "").upper()
        var_cells.add((sheet, cell))

    h = hashlib.sha1()
    # Version marker: changing the defined-name resolver must force a fresh
    # compile instead of loading a graph built without name rewriting.
    h.update(b"FASTSOLVER_PYCEL_FORMULA_REWRITE_V1\0")

    try:
        for name, target, sheet in sorted(_iter_defined_names(wb)):
            h.update(f"\x00NAME:{sheet or 'WORKBOOK'}!{name}={target}\x00".encode("utf-8"))
    except Exception:
        pass

    for ws in wb.worksheets:
        h.update(("\x00SHEET:" + ws.title + "\x00").encode("utf-8"))
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                if (ws.title, c.coordinate.upper()) in var_cells:
                    continue
                cv = c.value
                # openpyxl wraps array formulas in an ArrayFormula object
                # whose repr() embeds the Python id()/memory address, which
                # changes on every load. Hashing repr() therefore produced a
                # DIFFERENT digest every single run -> the cache could never
                # hit even on an identical, unchanged workbook. Hash the
                # array formula's stable text + range instead.
                if cv.__class__.__name__ == "ArrayFormula":
                    cv = f"ARRAYFORMULA({getattr(cv, 'ref', '')}:{getattr(cv, 'text', '')})"
                h.update(f"{c.coordinate}={cv!r};".encode("utf-8"))
    return h.hexdigest()[:16]


def _load_excel_cached(src_xlsx, variables, log_fn, objectives=None):
    """Compile a pycel-safe workbook with caching.

    Before pycel sees the workbook, expand supported defined names inside
    formula text to their cell/range targets and strip #REF! names. This
    prevents pycel 1.0b30 from treating normal names as missing tables.
    """
    src_xlsx = Path(src_xlsx)

    # Compute the cache key from the ORIGINAL workbook, BEFORE the expensive
    # formula-rewrite + defined-name sanitize. Those two passes touch ~2000+
    # formula cells and re-save the workbook every run (~2-3s) -- pure waste
    # on a cache hit. _compute_cache_key already hashes raw formula strings
    # and defined names (plus a version marker), so the original workbook
    # produces a stable, correct digest without needing the rewrite first.
    digest = _compute_cache_key(src_xlsx, variables, log_fn)

    # The cache MUST live in a location that survives between solves. The
    # previous path (src_xlsx.parent.parent) sat inside the disposable
    # per-run temp folder VBA creates each solve, so the .pkl was written
    # somewhere a future run never looks -> the cache could never hit and
    # every run re-paid the full pycel parse (~20s+). Use a stable, shared
    # cache directory keyed only by the workbook digest.
    import tempfile
    cache_dir = Path(tempfile.gettempdir()) / "fastsolver_cache"
    try:
        cache_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        cache_dir = Path(tempfile.gettempdir())
    # NOTE: bump this guard token whenever the formula-rewrite / defined-name
    # resolution logic changes. The digest hashes only the workbook, not this
    # module, so a stale pre-patch pickle would otherwise be reloaded and the
    # fix silently skipped. v6 = qualified-name workbook-scope fallback fix.
    cache_base = str(cache_dir / f"_pycel_cache_xmlnames_guard_v6_{digest}")
    cache_pkl = cache_base + ".pkl"

    def _obj_addrs():
        out = []
        for o in (objectives or []):
            if not o.get("active", True):
                continue
            ref = str(o.get("ref", "")).strip()
            if not ref or ref.startswith("="):
                continue
            try:
                out.append(_pycel_addr(ref))
            except Exception:
                pass
        return out

    def _graph_is_live(ex):
        """A pickle round-trip in pycel 1.0b30 can drop computed formula
        nodes, so every objective evaluates to None after reload. Detect that
        here and force a recompile instead of solving on a dead graph."""
        addrs = _obj_addrs()
        if not addrs:
            return True
        live = False
        for a in addrs:
            try:
                v = ex.evaluate(a)
            except Exception:
                continue
            if v is not None:
                live = True
                break
        return live

    if os.path.exists(cache_pkl):
        try:
            ex = ExcelCompiler.from_file(cache_pkl)
            name_maps = _build_name_maps_only(src_xlsx, log_fn)
            ex._fastsolver_name_maps = name_maps
            if _graph_is_live(ex):
                log_fn("Pycel graph loaded from cache (skipped parse + rewrite)")
                return ex
            log_fn("Cached graph reloaded dead (objectives=None); recompiling")
        except Exception as e:
            log_fn(f"Cache load failed ({e}); recompiling")

    # Cache miss: do the full (expensive) preprocessing path.
    compile_xlsx, name_maps = _rewrite_workbook_formulas_for_pycel(src_xlsx, log_fn)
    compile_xlsx = _sanitize_defined_names_for_pycel(compile_xlsx, log_fn)

    ex = ExcelCompiler(filename=str(compile_xlsx))
    ex._fastsolver_name_maps = name_maps

    # Warm the graph: evaluate every objective once BEFORE to_file(). pycel
    # 1.0b30 only serializes nodes it has actually computed; without this the
    # reloaded pickle returns None for every objective and the solver scores a
    # flat hard-penalty wall. Warming makes the cached graph reload live.
    try:
        for a in _obj_addrs():
            try:
                ex.evaluate(a)
            except Exception:
                pass
    except Exception:
        pass

    try:
        ex.to_file(cache_base, file_types=('pkl',))
        # Verify the just-written pickle actually reloads live before we trust
        # it on the next run. If it doesn't, drop it so the next run recompiles
        # rather than silently solving on a dead graph.
        try:
            chk = ExcelCompiler.from_file(cache_base + ".pkl")
            if not _graph_is_live(chk):
                os.remove(cache_base + ".pkl")
                log_fn("Cache self-check failed (dead reload); cache disabled this run")
            else:
                log_fn("Pycel graph compiled and cached for next run")
        except Exception:
            log_fn("Pycel graph compiled and cached for next run")
    except Exception as e:
        log_fn(f"Could not write pycel cache ({e}); continuing without it")
    return ex


def main(run_dir):
    raw_arg = run_dir
    run_dir = Path(run_dir)
    cfg_path = run_dir / "config.json"
    src_xlsx = run_dir / "source.xlsx"

    if not run_dir.is_dir() or not cfg_path.is_file():
        print("=" * 60, flush=True)
        print(" FASTSOLVER - STARTUP ERROR", flush=True)
        print("=" * 60, flush=True)
        print(f" The run directory is not valid:", flush=True)
        print(f"   argument received : {raw_arg!r}", flush=True)
        print(f"   resolved path     : {run_dir}", flush=True)
        print(f"   directory exists  : {run_dir.is_dir()}", flush=True)
        print(f"   config.json found : {cfg_path.is_file()}", flush=True)
        print(flush=True)
        if str(raw_arg).startswith("--"):
            print(" It looks like an OPTION ('--...') was passed as the run", flush=True)
            print(" directory. This almost always means the FastSolver.exe", flush=True)
            print(" being run is an OLD build that does not understand", flush=True)
            print(" monitor mode. Rebuild the exe from the current", flush=True)
            print(" 07_fastsolver_bridge.py (run 11_build_exe.py) and", flush=True)
            print(" redeploy the dist/FastSolver folder next to the .xlam.", flush=True)
        else:
            print(" The VBA side did not create config.json here, or the", flush=True)
            print(" path is wrong. Check modPythonBridge run-folder logic.", flush=True)
        print(flush=True)
        print(" This window is staying open. Type q then ENTER to close.",
              flush=True)
        _wait_for_q()
        sys.exit(3)

    log = []
    def log_line(s):
        log.append(s)
        print(s, flush=True)

    try:
        config = json.loads(cfg_path.read_text())
        log_line(f"Loaded config: {len(config['variables'])} vars, "
                 f"{len(config['objectives'])} objs, "
                 f"{len(config['constraints'])} cons")

        variables = config["variables"]
        n = len(variables)
        lo = np.array([float(v["lo"]) for v in variables])
        hi = np.array([float(v["hi"]) for v in variables])
        simplex_groups = collect_simplex_groups(variables)
        log_line(f"Simplex groups: {len(simplex_groups)}")

        _set_phase(run_dir, "compiling")
        t0 = time.time()
        excel = _load_excel_cached(src_xlsx, variables, log_line,
                                   objectives=config.get("objectives"))
        log_line(f"Pycel ready in {time.time()-t0:.2f}s")
        _set_phase(run_dir, "solving")

        x0 = np.zeros(n)
        for i, v in enumerate(variables):
            x0[i] = _safe_eval(excel, _pycel_addr(v["ref"]))
        x0 = apply_projection(x0, simplex_groups, lo, hi)

        score = build_score_fn(excel, config, variables, simplex_groups, lo, hi)
        score.log_fn = log_line
        score.run_dir = str(run_dir)
        start_score = score(x0)

        if os.environ.get("FASTSOLVER_DEBUG"):
            x_test = x0.copy()
            if len(x_test) > 0:
                x_test[0] = min(hi[0], x0[0] + 0.1) if x0[0] + 0.1 <= hi[0] else max(lo[0], x0[0] - 0.1)
            probe_score = score(x_test)
            log_line(f"Probe score (perturbed x[0]): {probe_score:.6f}")
            if abs(probe_score - start_score) < 1e-12:
                log_line("WARNING: score did not change on perturbation — pycel not propagating vars")
            x_zero = x0.copy()
            x_zero[0] = 0.0
            crash_score = score(x_zero)
            log_line(f"Score at x[0]=0 (div-by-zero test): {crash_score:.2e}")
        log_line(f"Initial score: {start_score:.6f}")

        method = config.get("method", "Auto")
        method_key = _method_key(method)
        max_iter = int(config.get("max_iter", 200))
        log_line(f"Method requested: {method!r} -> {method_key}; max_iter={max_iter}")

        t1 = time.time()
        if method_key in ("portfolio", "portfolio-cd", "portfolio-powell"):
            if method_key == "portfolio-powell":
                p_order = ("slsqp", "l-bfgs-b", "powell",
                           "differential evolution")
            else:
                p_order = ("slsqp", "l-bfgs-b",
                           "coordinate descent anti-zigzag adaptive step",
                           "differential evolution")
            x_best, f_best = portfolio_solve(
                score, x0, list(zip(lo, hi)), simplex_groups, lo, hi,
                max_iter=max(max_iter * 6, 1500), log_fn=log_line,
                order=p_order)
        elif method_key in ("cd-azas", "de-multi-restart"):
            x_best, f_best = run_method(method_key, score, x0,
                                        list(zip(lo, hi)),
                                        simplex_groups, lo, hi,
                                        max_iter)
        else:
            x_best, f_best = multi_start_solve(method_key, score, x0,
                                               list(zip(lo, hi)),
                                               simplex_groups, lo, hi,
                                               max_iter, n_starts=3)
        elapsed = time.time() - t1
        log_line(f"Optimized in {elapsed:.2f}s. Final score: {f_best:.6f}. "
                 f"Eval count: {score.eval_count[0]}")

        if os.environ.get("FASTSOLVER_DEBUG"):
            wb = load_workbook(src_xlsx)
            for v, val in zip(variables, x_best):
                sheet, cell = _split_addr(v["ref"])
                try:
                    wb[sheet][cell] = float(val)
                except Exception:
                    pass
            wb.save(run_dir / "solved.xlsx")

        results = {
            "status": "ok",
            "method_used": method,
            "elapsed_sec": elapsed,
            "eval_count": score.eval_count[0],
            "start_score": start_score,
            "final_score": float(f_best),
            "variable_names": [v["ref"] for v in variables],
            "variable_values": [float(x) for x in x_best],
            "log": log,
        }
        (run_dir / "results.json").write_text(json.dumps(results, indent=2))
        log_line("Wrote results.json + solved.xlsx")
    except Exception as e:
        tb = traceback.format_exc()
        wrote = False
        try:
            err = {
                "status": "error",
                "message": str(e),
                "traceback": tb,
                "log": log,
            }
            (run_dir / "results.json").write_text(json.dumps(err, indent=2))
            wrote = True
        except Exception as e2:
            print(f"(could not write results.json: {e2})", flush=True)

        print(flush=True)
        print("=" * 60, flush=True)
        print(" FASTSOLVER - SOLVE FAILED", flush=True)
        print("=" * 60, flush=True)
        print(f" Error: {e}", flush=True)
        print(flush=True)
        print(" --- traceback ---", flush=True)
        for ln in tb.splitlines()[-25:]:
            print(" | " + ln, flush=True)
        if log:
            print(flush=True)
            print(" --- log tail ---", flush=True)
            for ln in log[-15:]:
                print(" | " + str(ln), flush=True)
        if wrote:
            print("\n (results.json written - the Excel side will report "
                  "failure too)", flush=True)
        print(flush=True)
        print(" This window is staying open. Type q then ENTER to close.",
              flush=True)
        _wait_for_q()
        sys.exit(2)


def _wait_for_q():
    """Block until the user types q + Enter. Robust against a frozen-exe
    console where bare input() can return instantly."""
    import time as _t
    while True:
        try:
            ans = input(" > ").strip().lower()
        except Exception:
            _t.sleep(900.0)
            return
        if ans == "q":
            return
        print(" (type q then ENTER to close)", flush=True)


def _write_progress(run_dir, text):
    """Append a progress line the monitor window can display."""
    try:
        p = Path(run_dir) / "progress.txt"
        with open(p, "a", encoding="utf-8") as fh:
            fh.write(text.rstrip() + "\n")
    except Exception:
        pass


def _set_phase(run_dir, phase):
    """Signal the monitor which phase we're in: 'compiling' or 'solving'."""
    try:
        (Path(run_dir) / "phase.txt").write_text(phase, encoding="utf-8")
    except Exception:
        pass


def monitor(run_dir):
    """Visible companion window: animated compile bar, then live per-method
    solving counters. Reads phase.txt + progress.txt and self-closes when
    results.json appears."""
    import time as _t
    run_dir = Path(run_dir)
    prog = run_dir / "progress.txt"
    phase_f = run_dir / "phase.txt"
    done = run_dir / "results.json"

    def read_phase():
        try:
            return phase_f.read_text(encoding="utf-8").strip()
        except Exception:
            return "compiling"

    def clear():
        print("\r" + " " * 70 + "\r", end="")

    print("=" * 60)
    print("   F A S T S O L V E R")
    print("   (progress monitor - closes itself when finished)")
    print("=" * 60, flush=True)

    bar_w = 28
    pos = 0
    direction = 1
    comp_waited = 0.0
    COMPILE_LIMIT = 300.0
    print()
    while not done.exists() and read_phase() == "compiling":
        cells = ["."] * bar_w
        cells[pos] = "#"
        bar = "".join(cells)
        print(f"\r  Compiling workbook  [{bar}]  (this can take a few s)",
              end="", flush=True)
        pos += direction
        if pos >= bar_w - 1 or pos <= 0:
            direction *= -1
        _t.sleep(0.08)
        comp_waited += 0.08
        if comp_waited >= COMPILE_LIMIT:
            print(flush=True)
            print(f"  Still 'compiling' after {int(COMPILE_LIMIT)}s with no "
                  f"result - the solver likely crashed during startup.",
                  flush=True)
            break
    clear()
    if done.exists() or read_phase() != "compiling":
        print("\r  Compiling workbook  [ done ]", flush=True)

    print()
    print("  Solving:")
    seen = 0
    spin = "|/-\\"
    spin_i = 0
    waited = 0.0
    silent = 0.0
    SILENCE_LIMIT = 80.0
    while not done.exists():
        lines = []
        try:
            if prog.exists():
                lines = prog.read_text(encoding="utf-8").splitlines()
        except Exception:
            lines = []

        if len(lines) > seen:
            for ln in lines[seen:]:
                print("   " + ln, flush=True)
            seen = len(lines)
            silent = 0.0
        else:
            spin_i = (spin_i + 1) % len(spin)
            last = lines[-1] if lines else "starting..."
            print(f"\r   {spin[spin_i]} {last}   "
                  f"({waited:0.0f}s)        ",
                  end="", flush=True)
            silent += 0.4

        if silent >= SILENCE_LIMIT:
            print(flush=True)
            print(f"  No progress for {int(SILENCE_LIMIT)}s and no result "
                  f"- assuming the solver crashed.", flush=True)
            break

        _t.sleep(0.4)
        waited += 0.4

    print()
    print("  " + "-" * 50)

    failed = False
    msg = ""
    tb = ""
    logtail = []
    try:
        if done.exists():
            raw = done.read_text(encoding="utf-8", errors="replace")
            try:
                res = json.loads(raw)
            except Exception:
                failed = True
                msg = "results.json is present but not valid JSON " \
                      "(solver likely died mid-write)."
                tb = raw[-2000:]
                res = None
            if res is not None:
                if res.get("status") == "ok":
                    print(f"  DONE.  Final score: "
                          f"{res.get('final_score'):.6g}"
                          f"   ({res.get('eval_count','?')} evals)",
                          flush=True)
                else:
                    failed = True
                    msg = res.get("message", "(no message in results.json)")
                    tb = res.get("traceback", "") or ""
                    lg = res.get("log", [])
                    if isinstance(lg, list):
                        logtail = [str(x) for x in lg][-20:]
        else:
            failed = True
            msg = "Solver exited without producing results.json (it crashed)."
    except Exception as e:
        failed = True
        msg = f"Monitor could not read results.json: {e}"

    if not failed:
        print("  (window closes in 3s)", flush=True)
        _t.sleep(3.0)
        return

    print(flush=True)
    print("  " + "!" * 52, flush=True)
    print("  SOLVE FAILED", flush=True)
    print("  " + "!" * 52, flush=True)
    print(f"  Reason: {msg}", flush=True)

    if logtail:
        print("\n  --- solver log (last lines) ---", flush=True)
        for ln in logtail:
            print("  | " + ln, flush=True)

    if tb:
        print("\n  --- Python traceback ---", flush=True)
        for ln in tb.splitlines()[-25:]:
            print("  | " + ln, flush=True)

    print(flush=True)
    print("  " + "=" * 52, flush=True)
    print("  This window is STAYING OPEN so you can read the error.",
          flush=True)
    print("  Type  q  then ENTER to close it.", flush=True)
    print("  " + "=" * 52, flush=True)

    while True:
        try:
            ans = input("  > ").strip().lower()
        except Exception:
            _t.sleep(600.0)
            break
        if ans == "q":
            break
        print("  (type q then ENTER to close)", flush=True)


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "--monitor":
        monitor(sys.argv[2])
        sys.exit(0)
    if len(sys.argv) < 2:
        print("Usage: fastsolver_bridge.py <run_dir>", file=sys.stderr)
        print("       fastsolver_bridge.py --monitor <run_dir>",
              file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1])